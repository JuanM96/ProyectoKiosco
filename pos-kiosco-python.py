import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime, timedelta
import pandas as pd
from fpdf import FPDF
import os
import sys
import logging
from PIL import Image, ImageTk
import requests
import uuid
import hashlib
import json
from pathlib import Path

class LicenseManager:
    def __init__(self):
        self.config = self.load_config()
        self.firebase_url = "https://licenciaskioscopos-default-rtdb.firebaseio.com/"
        self.license_file = Path("license.json")
        self.machine_id = self.get_machine_id()
    
    def load_config(self):
        """Carga la configuración desde config.json"""
        try:
            # Buscar config.json en la carpeta de licencias
            config_paths = [
                Path("license_management/config.json"),
                Path("config.json")  # fallback
            ]
            
            for config_path in config_paths:
                if config_path.exists():
                    with open(config_path, 'r', encoding='utf-8') as f:
                        return json.load(f)
        except Exception as e:
            print(f"Error cargando configuración: {e}")
        
        # Configuración por defecto si no existe el archivo
        return {
            "firebase": {
                "url": "example.firebaseio.com",
            },
            "license": {
                "offline_tolerance_days": 7,
                "default_license_months": 1
            },
            "support": {
                "email": "juanignaciomurcia@gmail.com",   # Reemplazar con email real
                "phone": "1123909301"   # Reemplazar con teléfono real
            }
        }
    
    def get_machine_id(self):
        """Genera un ID único de la máquina basado en características del hardware"""
        try:
            # Combinamos varios identificadores únicos del sistema
            computer_name = os.environ.get('COMPUTERNAME', '')
            username = os.environ.get('USERNAME', '')
            
            # Creamos un hash único basado en estos valores
            unique_string = f"{computer_name}_{username}_{sys.platform}"
            machine_id = hashlib.sha256(unique_string.encode()).hexdigest()[:16]
            return machine_id
        except Exception as e:
            # Si hay algún error, generamos un ID aleatorio
            return str(uuid.uuid4())[:16]
    
    def save_license_locally(self, license_data):
        """Guarda los datos de licencia localmente"""
        try:
            with open(self.license_file, 'w') as f:
                json.dump(license_data, f)
        except Exception as e:
            print(f"Error guardando licencia: {e}")
    
    def load_license_locally(self):
        """Carga los datos de licencia guardados localmente"""
        try:
            if self.license_file.exists():
                with open(self.license_file, 'r') as f:
                    return json.load(f)
            return None
        except Exception as e:
            print(f"Error cargando licencia local: {e}")
            return None
    
    def check_and_update_expired_license(self, license_data):
        """Desactiva automáticamente licencias expiradas en Firebase"""
        if not license_data:
            return None
        
        try:
            expiry_date = datetime.fromisoformat(license_data.get('expiry_date', ''))
            current_date = datetime.now()
            
            # Si la licencia expiró Y sigue activa, desactivarla
            if current_date > expiry_date and license_data.get('active', False):
                print("🔄 Licencia expirada detectada, desactivando...")
                
                # Actualizar solo el campo 'active' en Firebase
                url = f"{self.firebase_url}/licenses/{self.machine_id}/active.json"
                response = requests.put(url, json=False, timeout=5)
                
                if response.status_code == 200:
                    # Actualizar datos locales
                    license_data['active'] = False
                    self.save_license_locally(license_data)
                    print("✅ Licencia desactivada automáticamente")
                else:
                    print(f"⚠️ No se pudo desactivar en Firebase: {response.status_code}")
                    
            return license_data
        except Exception as e:
            print(f"Error verificando expiración: {e}")
            return license_data

    def check_license_firebase(self):
        """Verifica la licencia en Firebase y actualiza si expiró"""
        try:
            url = f"{self.firebase_url}/licenses/{self.machine_id}.json"
            response = requests.get(url, timeout=10)
            
            if response.status_code == 200:
                license_data = response.json()
                if license_data:
                    # Verificar y actualizar si expiró
                    return self.check_and_update_expired_license(license_data)
            return None
        except Exception as e:
            print(f"Error conectando con Firebase: {e}")
            return None
    
    def is_license_valid(self, license_data):
        """Verifica si la licencia está vigente"""
        if not license_data:
            return False
        
        try:
            expiry_date = datetime.fromisoformat(license_data.get('expiry_date', ''))
            current_date = datetime.now()
            
            is_active = license_data.get('active', False)
            is_not_expired = current_date <= expiry_date
            
            return is_active and is_not_expired
        except Exception as e:
            print(f"Error validando licencia: {e}")
            return False
    
    def validate_license(self):
        """Método principal para validar la licencia"""
        # Primero intentamos verificar online
        online_license = self.check_license_firebase()
        
        if online_license:
            # Si encontramos licencia online, la guardamos localmente
            self.save_license_locally(online_license)
            return self.is_license_valid(online_license)
        
        # Si no hay conexión, verificamos la licencia local
        local_license = self.load_license_locally()
        if local_license:
            # Verificamos que no haya expirado hace más de 7 días (tolerancia)
            if self.is_license_valid_offline(local_license):
                return True
        
        return False
    
    def is_license_valid_offline(self, license_data):
        """Verifica la licencia offline con tolerancia configurable"""
        if not license_data:
            return False
        
        try:
            expiry_date = datetime.fromisoformat(license_data.get('expiry_date', ''))
            current_date = datetime.now()
            
            # Tolerancia configurable para validación offline
            tolerance_days = self.config['license']['offline_tolerance_days']
            tolerance = timedelta(days=tolerance_days)
            is_active = license_data.get('active', False)
            is_within_tolerance = current_date <= (expiry_date + tolerance)
            
            return is_active and is_within_tolerance
        except Exception as e:
            print(f"Error validando licencia offline: {e}")
            return False
    
    def show_license_error(self, root):
        """Muestra el diálogo de error de licencia"""
        error_window = tk.Toplevel(root)
        error_window.title("Error de Licencia")
        error_window.geometry("420x350")
        error_window.configure(bg="#F8F8F8")
        error_window.resizable(False, False)
        
        # Centrar la ventana
        error_window.transient(root)
        error_window.grab_set()
        
        # Icono de error
        try:
            error_window.iconbitmap("img/kiosco.ico")
        except:
            pass
        
        # Frame principal con scrollbar si es necesario
        main_frame = tk.Frame(error_window, bg="#F8F8F8")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        title_label = tk.Label(
            main_frame,
            text="⚠️ Licencia No Válida",
            font=("Arial", 16, "bold"),
            fg="#D32F2F",
            bg="#F8F8F8"
        )
        title_label.pack(pady=(0, 10))
        
        # Mensaje
        message_label = tk.Label(
            main_frame,
            text="Su licencia ha expirado o no es válida.\n\n"
                 "Para continuar usando el software,\n"
                 "contacte con soporte técnico.\n\n"
                 f"ID de Máquina: {self.machine_id}",
            font=("Arial", 10),
            fg="#333333",
            bg="#F8F8F8",
            justify="center"
        )
        message_label.pack(pady=(0, 15))
        
        # Información de contacto
        contact_info = self.config['support']
        contact_frame = tk.Frame(main_frame, bg="#E3F2FD", relief="solid", bd=1)
        contact_frame.pack(pady=(0, 15), padx=5, fill="x")
        
        contact_title = tk.Label(
            contact_frame,
            text="Para contactar al soporte:",
            font=("Arial", 10, "bold"),
            fg="#1976D2",
            bg="#E3F2FD"
        )
        contact_title.pack(pady=(8, 3))
        
        email_label = tk.Label(
            contact_frame,
            text=f"Email: {contact_info['email']}",
            font=("Arial", 9),
            fg="#333333",
            bg="#E3F2FD"
        )
        email_label.pack(pady=1)
        
        phone_label = tk.Label(
            contact_frame,
            text=f"Cel: {contact_info['phone']}",
            font=("Arial", 9),
            fg="#333333",
            bg="#E3F2FD"
        )
        phone_label.pack(pady=(1, 8))
        
        # Frame para el botón (para asegurar que aparezca)
        button_frame = tk.Frame(main_frame, bg="#F8F8F8", height=50)
        button_frame.pack(pady=(10, 0), fill="x")
        button_frame.pack_propagate(False)  # Mantener altura fija
        
        # Botón Aceptar (cerrar) - centrado
        accept_button = tk.Button(
            button_frame,
            text="Aceptar",
            font=("Arial", 11, "bold"),
            bg="#D32F2F",
            fg="white",
            padx=30,
            pady=10,
            cursor="hand2",
            command=lambda: self.close_application(root)
        )
        accept_button.place(relx=0.5, rely=0.5, anchor="center")  # Centrado absoluto
        
        # Manejar el cierre de ventana
        error_window.protocol("WM_DELETE_WINDOW", lambda: self.close_application(root))
    
    def close_application(self, root):
        """Cierra la aplicación"""
        root.quit()
        root.destroy()
        sys.exit()

class KioscoPOS:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema POS - Kiosco Argentina")
        
        # Maximizar ventana al iniciar
        self.root.state('zoomed')  # Para Windows
        # Alternativa para otros sistemas: self.root.attributes('-zoomed', True)
        
        self.root.configure(bg="#FAF2E3")
        
        # Configurar ícono de manera segura
        try:
            icon_path = self.get_resource_path("img", "kiosco.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception as e:
            print(f"No se pudo cargar el ícono: {e}")
            # Continúa sin ícono si hay error
        
        # Verificar licencia antes de continuar
        self.license_manager = LicenseManager()
        if not self.verify_license():
            return  # Si no hay licencia válida, no continúa
        
        # Usuario actual
        self.usuario_actual = None
        
        # Inicializar base de datos
        self.init_database()
        
        # Carrito de compras
        self.carrito = []
        # Atajos globales habilitados por defecto (se deshabilitan mientras hay diálogos modales abiertos)
        self.shortcuts_enabled = True
        
        # Mostrar login
        self.mostrar_login()
        logging.basicConfig(level=logging.INFO, format='%(message)s')
    
    def verify_license(self):
        """Verifica la licencia antes de mostrar la aplicación"""
        if self.license_manager.validate_license():
            return True
        else:
            # Mostrar ventana de error y cerrar aplicación
            self.license_manager.show_license_error(self.root)
            return False
        
    def get_resource_path(self, *args):
        """Obtiene la ruta correcta para recursos tanto en desarrollo como en ejecutable"""
        try:
            # Cuando se ejecuta desde PyInstaller
            base_path = sys._MEIPASS
        except AttributeError:
            # Cuando se ejecuta desde el script normal
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        return os.path.join(base_path, *args)
        
    def init_database(self):
        """Inicializa la base de datos y crea las tablas"""
        self.conn = sqlite3.connect('kiosco.db')
        self.cursor = self.conn.cursor()
        
        # Tabla de usuarios
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                rol TEXT NOT NULL
            )
        ''')
        
        # Tabla de productos
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                precio REAL NOT NULL,
                costo REAL NOT NULL,
                stock INTEGER NOT NULL,
                categoria TEXT,
                codigo_barras TEXT
            )
        ''')
        
        # Tabla de ventas
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS ventas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                usuario TEXT NOT NULL,
                metodo_pago TEXT NOT NULL,
                total REAL NOT NULL,
                costo_total REAL NOT NULL,
                turno TEXT NOT NULL
            )
        ''')
        
        # Tabla de items de venta
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS items_venta (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                venta_id INTEGER NOT NULL,
                producto_nombre TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                precio_unitario REAL NOT NULL,
                costo_unitario REAL NOT NULL,
                FOREIGN KEY (venta_id) REFERENCES ventas (id)
            )
        ''')
        
        # Tabla de configuración
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS configuracion (
                clave TEXT PRIMARY KEY,
                valor TEXT NOT NULL
            )
        ''')
        
        # Insertar usuario admin por defecto si no existe
        self.cursor.execute("SELECT * FROM usuarios WHERE nombre = 'Administrador'")
        if not self.cursor.fetchone():
            self.cursor.execute('''
                INSERT INTO usuarios (nombre, password, rol) 
                VALUES ('Administrador', 'admin123', 'admin')
            ''')
        
        # Insertar productos de ejemplo si no existen
        self.cursor.execute("SELECT COUNT(*) FROM productos")
        if self.cursor.fetchone()[0] == 0:
            productos_ejemplo = [
                ('Coca Cola 500ml', 800, 500, 50, 'Bebidas', '7790895602764'),
                ('Alfajor Milka', 600, 400, 30, 'Golosinas', '7622210805164'),
                ('Papas Lays', 950, 650, 25, 'Snacks', '7790310510186')
            ]
            self.cursor.executemany('''
                INSERT INTO productos (nombre, precio, costo, stock, categoria, codigo_barras)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', productos_ejemplo)
        
        # Insertar configuración por defecto
        self.cursor.execute("SELECT * FROM configuracion WHERE clave = 'stock_habilitado'")
        if not self.cursor.fetchone():
            self.cursor.execute('''
                INSERT INTO configuracion (clave, valor) 
                VALUES ('stock_habilitado', '1')
            ''')
        
        self.conn.commit()
        
        # Limpiar códigos de barras con .0 al final
        self.limpiar_codigos_barras()

        # Ejecutar migraciones de BD
        self._migrar_bd()

    def _migrar_bd(self):
        """Ejecuta migraciones necesarias para actualizar el esquema de BD."""
        try:
            # Verificar y agregar columna precio_sugerido a productos
            self.cursor.execute("PRAGMA table_info(productos)")
            columnas = [col[1] for col in self.cursor.fetchall()]
            
            if 'precio_sugerido' not in columnas:
                # Agregar columna con valor calculado por defecto (usando 30% de ganancia)
                self.cursor.execute('''
                    ALTER TABLE productos ADD COLUMN precio_sugerido REAL DEFAULT 0
                ''')
                # Calcular precio sugerido para productos existentes (ganancia = 30%)
                self.cursor.execute('''
                    UPDATE productos SET precio_sugerido = 
                        CASE WHEN costo > 0 THEN costo / (1 - 0.30) ELSE 0 END
                ''')
                print("✅ Columna 'precio_sugerido' agregada a tabla 'productos'")
            
            if 'ganancia_deseada' not in columnas:
                # Agregar columna para % de ganancia deseada por producto (NULL = usar global)
                self.cursor.execute('''
                    ALTER TABLE productos ADD COLUMN ganancia_deseada REAL DEFAULT NULL
                ''')
                print("✅ Columna 'ganancia_deseada' agregada a tabla 'productos'")
            
            # Asegurar que existe la configuración global de ganancia deseada
            self.cursor.execute("SELECT valor FROM configuracion WHERE clave = 'ganancia_deseada_default'")
            if not self.cursor.fetchone():
                self.cursor.execute('''
                    INSERT INTO configuracion (clave, valor) VALUES ('ganancia_deseada_default', '30')
                ''')
                print("✅ Configuración 'ganancia_deseada_default' establecida a 30%")
            
            self.conn.commit()
        except Exception as e:
            print(f"⚠️ Error en migración de BD: {e}")
            self.conn.rollback()
    
    def get_configuracion(self, clave, default='1'):
        """Obtiene un valor de configuración"""
        self.cursor.execute("SELECT valor FROM configuracion WHERE clave = ?", (clave,))
        result = self.cursor.fetchone()
        return result[0] if result else default
    
    def set_configuracion(self, clave, valor):
        """Establece un valor de configuración"""
        self.cursor.execute('''
            INSERT OR REPLACE INTO configuracion (clave, valor)
            VALUES (?, ?)
        ''', (clave, valor))
        self.conn.commit()
    
    def stock_habilitado(self):
        """Verifica si el stock está habilitado"""
        return self.get_configuracion('stock_habilitado') == '1'
    
    def calcular_precio_sugerido(self, costo, ganancia_deseada=None):
        """Calcula el precio sugerido usando la fórmula: P = C / (1 - R)
        Donde P = Precio, C = Costo, R = Ganancia deseada (en decimal, ej: 0.30 para 30%)
        
        Si ganancia_deseada es None, usa el valor global.
        """
        if ganancia_deseada is None:
            # Obtener el % de ganancia deseada global
            ganancia_str = self.get_configuracion('ganancia_deseada_default', '30')
            try:
                ganancia_deseada = float(ganancia_str) / 100.0  # Convertir de % a decimal
            except ValueError:
                ganancia_deseada = 0.30  # Default fallback
        else:
            # Si se pasa el % como número, convertir a decimal si es necesario
            if ganancia_deseada > 1:
                ganancia_deseada = ganancia_deseada / 100.0
        
        # Evitar división por cero
        if ganancia_deseada >= 1.0:
            return 0.0
        
        if costo <= 0:
            return 0.0
        
        # Calcular precio sugerido y redondear a 1 decimal
        precio_sugerido = costo / (1.0 - ganancia_deseada)
        return round(precio_sugerido, 1)
    
    def mostrar_login(self):
        """Muestra la ventana de login"""
        self.login_frame = tk.Frame(self.root, bg='#FAF2E3')
        self.login_frame.place(relx=0.5, rely=0.5, anchor='center')

        try:
            ruta_logo = self.get_resource_path("img", "kioscoimg.png")
            imagen_logo = Image.open(ruta_logo)
            imagen_logo = imagen_logo.resize((350, 350))  # tamaño ajustable
            self.logo_img = ImageTk.PhotoImage(imagen_logo)
            tk.Label(
                self.login_frame,
                image=self.logo_img,
                bg='#FAF2E3'
            ).pack(pady=5)
        except Exception as e:
            print("No se pudo cargar el logo:", e)

        # Logo/Título
        # tk.Label(
        #     self.login_frame, 
        #     text="Sistema POS - Kiosco", 
        #     font=('Arial', 24, 'bold'),
        #     bg='#FAF2E3',
        #     fg='#2563eb'
        # ).pack(pady=20)
        
        # Usuario
        tk.Label(
            self.login_frame, 
            text="Usuario:", 
            font=('Arial', 12),
            bg='#FAF2E3'
        ).pack(pady=5)
        
        self.entry_usuario = tk.Entry(self.login_frame, font=('Arial', 12), width=25)
        self.entry_usuario.pack(pady=5)
        
        # Contraseña
        tk.Label(
            self.login_frame, 
            text="Contraseña:", 
            font=('Arial', 12),
            bg='#FAF2E3'
        ).pack(pady=5)
        
        self.entry_password = tk.Entry(self.login_frame, font=('Arial', 12), width=25, show='*')
        self.entry_password.pack(pady=5)
        self.entry_password.bind('<Return>', lambda e: self.login())
        
        # Botón login
        tk.Button(
            self.login_frame,
            text="Iniciar Sesión",
            font=('Arial', 12, 'bold'),
            bg='#2563eb',
            fg='white',
            command=self.login,
            width=20,
            cursor='hand2'
        ).pack(pady=20)
        
        # Texto ayuda
        # tk.Label(
        #     self.login_frame,
        #     text="Usuario demo: Administrador / admin123",
        #     font=('Arial', 9),
        #     bg='#FAF2E3',
        #     fg='gray'
        # ).pack()
    
    def login(self):
        """Procesa el login del usuario"""
        usuario = self.entry_usuario.get()
        password = self.entry_password.get()
        
        self.cursor.execute(
            "SELECT * FROM usuarios WHERE nombre = ? AND password = ?",
            (usuario, password)
        )
        user = self.cursor.fetchone()
        
        if user:
            self.usuario_actual = {
                'id': user[0],
                'nombre': user[1],
                'rol': user[3]
            }
            self.login_frame.destroy()
            self.mostrar_interfaz_principal()
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")
    
    def mostrar_interfaz_principal(self):
        """Muestra la interfaz principal del sistema"""
        # Header
        header = tk.Frame(self.root, bg='#D94A2B', height=60)
        header.pack(fill='x')
        
        tk.Label(
            header,
            text="Sistema POS - Kiosco",
            font=('Arial', 18, 'bold'),
            bg='#D94A2B',
            fg='white'
        ).pack(side='left', padx=20, pady=10)
        
        tk.Label(
            header,
            text=f"👤 {self.usuario_actual['nombre']}",
            font=('Arial', 11),
            bg='#D94A2B',
            fg='white'
        ).pack(side='right', padx=10)
        
        tk.Button(
            header,
            text="Cerrar Sesión",
            font=('Arial', 10),
            bg="#d10909",
            fg='white',
            command=self.logout,
            cursor='hand2'
        ).pack(side='right', padx=10)
        
        # Notebook (pestañas)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        self.notebook.style = ttk.Style()
        self.notebook.style.configure('TNotebook', background='#FAF2E3')
        
        # Crear pestañas
        self.turno_actual = None
        if self.usuario_actual['rol'] == 'empleado':
            self.seleccionar_turno()
        else:
            self.crear_pestaña_venta()
            self.crear_pestaña_usuarios()
            self.crear_pestaña_productos()
            self.crear_pestaña_reportes()
    
    def crear_pestaña_venta(self):
        """Crea la pestaña de punto de venta"""
        frame_venta = tk.Frame(self.notebook, bg='#FAF2E3')
        self.notebook.add(frame_venta, text='🛒 Punto de Venta')

        # Frame izquierdo - Búsqueda de productos
        frame_izq = tk.Frame(frame_venta, bg='#FAF2E3')
        frame_izq.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        
        tk.Label(
            frame_izq,
            text="Buscar Productos",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=5)
        
        # Búsqueda por nombre
        tk.Label(frame_izq, text="Buscar por nombre o categoría:", bg='#FAF2E3').pack(pady=5)
        self.entry_buscar = tk.Entry(frame_izq, font=('Arial', 11), width=40)
        self.entry_buscar.pack(pady=5)
        self.entry_buscar.bind('<KeyRelease>', lambda e: self.actualizar_lista_productos())
        
        # Búsqueda por código de barras
        tk.Label(frame_izq, text="Código de barras (presiona Enter):", bg='#FAF2E3').pack(pady=5)
        self.entry_barcode = tk.Entry(frame_izq, font=('Arial', 11), width=40)
        self.entry_barcode.pack(pady=5)
        self.entry_barcode.bind('<Return>', self.buscar_por_barcode)
        
        # Lista de productos
        frame_lista = tk.Frame(frame_izq, bg='#FAF2E3')
        frame_lista.pack(fill='both', expand=True, pady=10)
        
        scrollbar = tk.Scrollbar(frame_lista)
        scrollbar.pack(side='right', fill='y')
        
        self.lista_productos = tk.Listbox(
            frame_lista,
            font=('Arial', 10),
            yscrollcommand=scrollbar.set,
            height=20
        )
        self.lista_productos.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.lista_productos.yview)
        
        self.lista_productos.bind('<Double-Button-1>', lambda e: self.agregar_al_carrito())
        
        tk.Button(
            frame_izq,
            text="Agregar al Carrito",
            font=('Arial', 11, 'bold'),
            bg='#2563eb',
            fg='white',
            command=self.agregar_al_carrito,
            cursor='hand2'
        ).pack(pady=10)
        
        # Frame derecho - Carrito
        frame_der = tk.Frame(frame_venta, bg='#FAF2E3', width=400)
        frame_der.pack(side='right', fill='both', padx=10, pady=10)
        frame_der.pack_propagate(False)
        
        tk.Label(
            frame_der,
            text="Carrito de Venta",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=5)
        
        # Lista del carrito
        frame_carrito = tk.Frame(frame_der, bg='#FAF2E3')
        frame_carrito.pack(fill='both', expand=True, pady=10)
        
        scrollbar_carrito = tk.Scrollbar(frame_carrito)
        scrollbar_carrito.pack(side='right', fill='y')
        
        self.lista_carrito = tk.Listbox(
            frame_carrito,
            font=('Arial', 10),
            yscrollcommand=scrollbar_carrito.set
        )
        self.lista_carrito.pack(side='left', fill='both', expand=True)
        scrollbar_carrito.config(command=self.lista_carrito.yview)
        
        # Botones de carrito
        frame_botones = tk.Frame(frame_der, bg='#FAF2E3')
        frame_botones.pack(fill='x', pady=5)
        
        tk.Button(
            frame_botones,
            text="Eliminar",
            font=('Arial', 10),
            bg='#dc2626',
            fg='white',
            command=self.eliminar_del_carrito,
            cursor='hand2'
        ).pack(side='left', padx=5)

        tk.Button(
            frame_botones,
            text="Eliminar Unidad",
            font=('Arial', 10),
            bg='#dc2626',
            fg='white',
            command=self.eliminar_uno_del_carrito,
            cursor='hand2'
        ).pack(side='left', padx=5)

        tk.Button(
            frame_botones,
            text="Vaciar Carrito",
            font=('Arial', 10),
            bg='#ef4444',
            fg='white',
            command=self.vaciar_carrito,
            cursor='hand2'
        ).pack(side='left', padx=5)

        # Botón para restar stock (solo si está habilitado)
        self.btn_restar_stock = tk.Button(
            frame_botones,
            text="Restar Stock",
            font=('Arial', 10),
            bg='#f59e0b',
            fg='white',
            command=self.restar_stock_interno,
            cursor='hand2'
        )
        if self.stock_habilitado():
            self.btn_restar_stock.pack(side='left', padx=5)

        # Total
        self.label_total = tk.Label(
            frame_der,
            text="TOTAL: $0",
            font=('Arial', 18, 'bold'),
            bg='#FAF2E3',
            fg='#2563eb'
        )
        self.label_total.pack(pady=10)
        
        # Métodos de pago
        tk.Label(
            frame_der,
            text="Método de Pago:",
            font=('Arial', 11, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=5)
        
        metodos = ['Efectivo', 'Transferencia', 'Débito', 'Crédito']
        keys = ['F1', 'F2', 'F3', 'F4']
        for i, metodo in enumerate(metodos):
            key = keys[i] if i < len(keys) else ''
            btn_text = f"{key}  Cobrar - {metodo}"
            btn = tk.Button(
                frame_der,
                text=btn_text,
                font=('Arial', 11, 'bold'),
                bg='#16a34a',
                fg='white',
                command=lambda m=metodo: self.cobrar_metodo(m),
                cursor='hand2'
            )
            btn.pack(fill='x', pady=2)
            # Bind keyboard shortcut to the root window via central handler
            if key:
                try:
                    # Use lambda with default arg to avoid late binding
                    self.root.bind(f'<{key}>', lambda e, m=metodo: self._handle_global_shortcut(e, m))
                except Exception:
                    pass
        
        # Cargar productos
        self.actualizar_lista_productos()

    def _handle_global_shortcut(self, event, metodo):
        """Handler central para atajos de teclado (F1-F4).
        Respeta `self.shortcuts_enabled` para evitar que atajos globales
        se ejecuten mientras hay un diálogo modal abierto.
        """
        # Si los atajos están deshabilitados, no consumir el evento
        # así el diálogo modal (si existe) puede recibir la tecla.
        if not getattr(self, 'shortcuts_enabled', True):
            return None

        try:
            # Ejecutar el método de cobro asociado
            self.cobrar_metodo(metodo)
        except Exception:
            pass
        # Consumir la tecla para evitar dobles ejecuciones
        return 'break'
    
    def crear_pestaña_productos(self):
        """Crea la pestaña de gestión de productos"""
        frame_productos = tk.Frame(self.notebook, bg='#FAF2E3')
        self.notebook.add(frame_productos, text='📦 Productos')
        
        # Frame superior - Configuración
        frame_config = tk.Frame(frame_productos, bg='#E3F2FD', relief='raised', bd=2)
        frame_config.pack(fill='x', padx=10, pady=5)
        
        tk.Label(
            frame_config,
            text="⚙️ Configuración de Stock",
            font=('Arial', 12, 'bold'),
            bg='#E3F2FD'
        ).pack(side='left', padx=10, pady=5)
        
        # Variable para el checkbox de stock
        self.stock_var = tk.BooleanVar()
        self.stock_var.set(self.stock_habilitado())
        
        self.stock_checkbox = tk.Checkbutton(
            frame_config,
            text="Habilitar control de stock",
            variable=self.stock_var,
            font=('Arial', 10),
            bg='#E3F2FD',
            command=self.confirmar_toggle_stock
        )
        self.stock_checkbox.pack(side='left', padx=10, pady=5)
        
        tk.Label(
            frame_config,
            text="ℹ️ Desactivar esta opción eliminará las validaciones y descuentos de stock",
            font=('Arial', 9),
            bg='#E3F2FD',
            fg='gray'
        ).pack(side='left', padx=10, pady=5)
        
        # Separador visual
        ttk.Separator(frame_config, orient='vertical').pack(side='left', fill='y', padx=10)
        
        # Configuración de ganancia deseada
        tk.Label(
            frame_config,
            text="Ganancia Deseada (%):",
            font=('Arial', 10),
            bg='#E3F2FD'
        ).pack(side='left', padx=5)
        
        ganancia_default = self.get_configuracion('ganancia_deseada_default', '30')
        self.ganancia_deseada_var = tk.StringVar(value=ganancia_default)
        
        self.spinbox_ganancia_deseada = tk.Spinbox(
            frame_config,
            from_=1,
            to=99,
            textvariable=self.ganancia_deseada_var,
            font=('Arial', 10),
            width=5
        )
        self.spinbox_ganancia_deseada.pack(side='left', padx=5)
        
        tk.Button(
            frame_config,
            text="Guardar",
            font=('Arial', 9),
            bg='#16a34a',
            fg='white',
            command=self.guardar_ganancia_deseada_global,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        # Frame contenedor para formulario y tabla
        frame_contenedor = tk.Frame(frame_productos, bg='#FAF2E3')
        frame_contenedor.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Frame izquierdo - Formulario
        frame_form = tk.Frame(frame_contenedor, bg='#FAF2E3', width=350)
        frame_form.pack(side='left', fill='y', padx=10, pady=10)
        frame_form.pack_propagate(False)
        
        tk.Label(
            frame_form,
            text="Agregar/Editar Producto",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=10)
        
        # Campos del formulario
        self.producto_id = None
        
        tk.Label(frame_form, text="Nombre:", bg='#FAF2E3').pack(pady=2)
        self.prod_nombre = tk.Entry(frame_form, font=('Arial', 11), width=30)
        self.prod_nombre.pack(pady=2)
        
        tk.Label(frame_form, text="Precio de Venta ($):", bg='#FAF2E3').pack(pady=2)
        self.prod_precio = tk.Entry(frame_form, font=('Arial', 11), width=30)
        self.prod_precio.pack(pady=2)
        
        tk.Label(frame_form, text="Costo ($):", bg='#FAF2E3').pack(pady=2)
        self.prod_costo = tk.Entry(frame_form, font=('Arial', 11), width=30)
        self.prod_costo.pack(pady=2)
        
        # Referencias para poder ocultar/mostrar los campos de stock
        self.stock_label = tk.Label(frame_form, text="Stock:", bg='#FAF2E3')
        self.stock_label.pack(pady=2)
        self.prod_stock = tk.Entry(frame_form, font=('Arial', 11), width=30)
        self.prod_stock.pack(pady=2)
        
        tk.Label(frame_form, text="Categoría:", bg='#FAF2E3').pack(pady=2)
        self.prod_categoria = ttk.Combobox(
            frame_form,
            font=('Arial', 11),
            width=28,
            values=['Bebidas', 'Golosinas', 'Snacks', 'Cigarrillos', 'Lácteos', 'Panificados', 'Limpieza', 'Helados', 'Otros']
        )
        self.prod_categoria.pack(pady=2)
        
        tk.Label(frame_form, text="Código de Barras:", bg='#FAF2E3').pack(pady=2)
        self.prod_barcode = tk.Entry(frame_form, font=('Arial', 11), width=30)
        self.prod_barcode.pack(pady=2)
        
        tk.Label(frame_form, text="Ganancia Deseada (%) [vacío=global]:", bg='#FAF2E3').pack(pady=2)
        self.prod_ganancia_deseada = tk.Entry(frame_form, font=('Arial', 11), width=30)
        self.prod_ganancia_deseada.pack(pady=2)
        
        # Botones
        frame_botones_prod = tk.Frame(frame_form, bg='#FAF2E3')
        frame_botones_prod.pack(pady=20)
        
        tk.Button(
            frame_botones_prod,
            text="Guardar",
            font=('Arial', 11, 'bold'),
            bg='#2563eb',
            fg='white',
            command=self.guardar_producto,
            width=12,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Button(
            frame_botones_prod,
            text="Limpiar",
            font=('Arial', 11),
            bg='#6b7280',
            fg='white',
            command=self.limpiar_formulario_producto,
            width=12,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        # Frame derecho - Lista de productos
        frame_lista_prod = tk.Frame(frame_contenedor, bg='#FAF2E3')
        frame_lista_prod.pack(side='right', fill='both', expand=True, padx=10, pady=10)
        
        tk.Label(
            frame_lista_prod,
            text="Inventario de Productos",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=10)

        # Campo de búsqueda para productos (filtra la tabla de productos)
        search_frame = tk.Frame(frame_lista_prod, bg='#FAF2E3')
        search_frame.pack(fill='x', padx=5, pady=(0, 8))
        tk.Label(search_frame, text="Buscar:", font=('Arial', 10), bg='#FAF2E3').pack(side='left', padx=(0, 6))
        self.entry_buscar_productos = tk.Entry(search_frame, font=('Arial', 10), width=40)
        self.entry_buscar_productos.pack(side='left', fill='x', expand=True)
        self.entry_buscar_productos.bind('<KeyRelease>', lambda e: self.actualizar_tabla_productos())
        
        # Tabla de productos
        frame_tabla = tk.Frame(frame_lista_prod, bg='#FAF2E3')
        frame_tabla.pack(fill='both', expand=True)
        
        scrollbar_tabla = tk.Scrollbar(frame_tabla)
        scrollbar_tabla.pack(side='right', fill='y')
        
        # Columnas condicionadas por configuración de stock
        if self.stock_habilitado():
            columnas = ('ID', 'Nombre', 'Precio', 'Costo', 'Stock', 'Sugerido', 'Categoría', 'Código')
        else:
            columnas = ('ID', 'Nombre', 'Precio', 'Costo', 'Sugerido', 'Categoría', 'Código')
        
        self.tabla_productos = ttk.Treeview(
            frame_tabla,
            columns=columnas,
            show='headings',
            yscrollcommand=scrollbar_tabla.set
        )
        
        self.tabla_productos.heading('ID', text='ID', command=lambda: self.ordenar_tabla_productos('ID'))
        self.tabla_productos.heading('Nombre', text='Nombre', command=lambda: self.ordenar_tabla_productos('Nombre'))
        self.tabla_productos.heading('Precio', text='Precio', command=lambda: self.ordenar_tabla_productos('Precio'))
        self.tabla_productos.heading('Costo', text='Costo', command=lambda: self.ordenar_tabla_productos('Costo'))
        if self.stock_habilitado():
            self.tabla_productos.heading('Stock', text='Stock', command=lambda: self.ordenar_tabla_productos('Stock'))
        self.tabla_productos.heading('Sugerido', text='Sugerido', command=lambda: self.ordenar_tabla_productos('Sugerido'))
        self.tabla_productos.heading('Categoría', text='Categoría', command=lambda: self.ordenar_tabla_productos('Categoría'))
        self.tabla_productos.heading('Código', text='Código Barras', command=lambda: self.ordenar_tabla_productos('Código'))
            
        self.tabla_productos.column('ID', width=50)
        self.tabla_productos.column('Nombre', width=200)
        self.tabla_productos.column('Precio', width=80)
        self.tabla_productos.column('Costo', width=80)
        if self.stock_habilitado():
            self.tabla_productos.column('Stock', width=60)
        self.tabla_productos.column('Sugerido', width=80)
        self.tabla_productos.column('Categoría', width=100)
        self.tabla_productos.column('Código', width=120)
        
        self.tabla_productos.pack(side='left', fill='both', expand=True)
        scrollbar_tabla.config(command=self.tabla_productos.yview)
        
        self.tabla_productos.bind('<Double-Button-1>', self.editar_producto)
        
        # Botones de acción
        frame_acciones = tk.Frame(frame_lista_prod, bg='#FAF2E3')
        frame_acciones.pack(fill='x', pady=10)
        
        tk.Button(
            frame_acciones,
            text="Editar Seleccionado",
            font=('Arial', 10),
            bg='#ea580c',
            fg='white',
            command=self.editar_producto,
            cursor='hand2'
        ).pack(side='left', padx=5)
        tk.Button(
            frame_acciones,
            text="Duplicar Seleccionado",
            font=('Arial', 10),
            bg='#0ea5e9',
            fg='white',
            command=self.duplicar_producto,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Button(
            frame_acciones,
            text="Eliminar Seleccionado",
            font=('Arial', 10),
            bg='#dc2626',
            fg='white',
            command=self.eliminar_producto,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Button(
            frame_acciones,
            text="Exportar a Excel",
            font=('Arial', 10),
            bg='#16a34a',
            fg='white',
            command=self.exportar_productos_excel,
            cursor='hand2'
        ).pack(side='right', padx=5)
        tk.Button(
        frame_acciones,
        text="Importar Productos",
        font=('Arial', 10),
        bg='#0ea5e9',
        fg='white',
        command=self.importar_productos,
        cursor='hand2'
        ).pack(side='right', padx=5)
        
        # Separador visual antes del botón peligroso
        tk.Label(
            frame_acciones,
            text=" | ",
            font=('Arial', 12, 'bold'),
            bg='#FAF2E3',
            fg='gray'
        ).pack(side='right', padx=10)
        
        # Botón para eliminar todos los productos (peligroso)
        tk.Button(
            frame_acciones,
            text="ELIMINAR TODO",
            font=('Arial', 10, 'bold'),
            bg='#dc2626',
            fg='white',
            command=self.confirmar_eliminar_productos,
            cursor='hand2',
            relief='raised',
            bd=3,
            width=16
        ).pack(side='right', padx=10)
        # Cargar productos
        self.actualizar_tabla_productos()
        
        # Actualizar visibilidad de campos de stock
        self.actualizar_visibilidad_stock()
    
    def confirmar_toggle_stock(self):
        """Confirma el cambio de configuración de stock antes de aplicarlo"""
        nuevo_estado = self.stock_var.get()
        estado_actual = self.stock_habilitado()
        
        # Si no hay cambio real, no hacer nada
        if (nuevo_estado and estado_actual) or (not nuevo_estado and not estado_actual):
            return
        
        if nuevo_estado:
            # Activando stock
            mensaje = ("⚠️ ACTIVAR CONTROL DE STOCK\n\n"
                      "Esta acción habilitará:\n"
                      "• Validación de stock en ventas\n"
                      "• Descuento automático de stock\n"
                      "• Campos y columnas de stock visibles\n"
                      "• Alertas de stock bajo\n\n"
                      "¿Confirmas activar el control de stock?")
            titulo = "Confirmar Activación de Stock"
        else:
            # Desactivando stock
            mensaje = ("⚠️ DESACTIVAR CONTROL DE STOCK\n\n"
                      "Esta acción deshabilitará:\n"
                      "• Validación de stock en ventas\n"
                      "• Descuento automático de stock\n"
                      "• Campos y columnas de stock se ocultarán\n"
                      "• No habrá alertas de stock bajo\n\n"
                      "¿Confirmas desactivar el control de stock?")
            titulo = "Confirmar Desactivación de Stock"
        
        # Mostrar diálogo de confirmación
        if messagebox.askyesno(titulo, mensaje, icon='warning'):
            # Usuario confirmó el cambio
            self.toggle_stock()
        else:
            # Usuario canceló, revertir el checkbox
            self.stock_var.set(estado_actual)
    
    def toggle_stock(self):
        """Cambia la configuración de stock habilitado/deshabilitado"""
        nuevo_valor = '1' if self.stock_var.get() else '0'
        self.set_configuracion('stock_habilitado', nuevo_valor)
        
        # Actualizar visibilidad de los campos
        self.actualizar_visibilidad_stock()
        
        # Actualizar visibilidad del botón "Restar Stock" si existe
        if hasattr(self, 'btn_restar_stock'):
            if self.stock_habilitado():
                self.btn_restar_stock.pack(side='left', padx=5)
            else:
                self.btn_restar_stock.pack_forget()
        
        # Recrear la tabla de productos con las columnas correctas
        if hasattr(self, 'tabla_productos'):
            # Guardar el frame padre
            frame_padre = self.tabla_productos.master
            
            # Destruir tabla anterior
            self.tabla_productos.destroy()
            
            # Recrear tabla con columnas correctas
            if self.stock_habilitado():
                columnas = ('ID', 'Nombre', 'Precio', 'Costo', 'Stock', 'Sugerido', 'Categoría', 'Código')
            else:
                columnas = ('ID', 'Nombre', 'Precio', 'Costo', 'Sugerido', 'Categoría', 'Código')
            
            self.tabla_productos = ttk.Treeview(
                frame_padre,
                columns=columnas,
                show='headings',
                yscrollcommand=frame_padre.children['!scrollbar'].set
            )
            
            # Configurar encabezados
            self.tabla_productos.heading('ID', text='ID', command=lambda: self.ordenar_tabla_productos('ID'))
            self.tabla_productos.heading('Nombre', text='Nombre', command=lambda: self.ordenar_tabla_productos('Nombre'))
            self.tabla_productos.heading('Precio', text='Precio', command=lambda: self.ordenar_tabla_productos('Precio'))
            self.tabla_productos.heading('Costo', text='Costo', command=lambda: self.ordenar_tabla_productos('Costo'))
            if self.stock_habilitado():
                self.tabla_productos.heading('Stock', text='Stock', command=lambda: self.ordenar_tabla_productos('Stock'))
            self.tabla_productos.heading('Sugerido', text='Sugerido', command=lambda: self.ordenar_tabla_productos('Sugerido'))
            self.tabla_productos.heading('Categoría', text='Categoría', command=lambda: self.ordenar_tabla_productos('Categoría'))
            self.tabla_productos.heading('Código', text='Código Barras', command=lambda: self.ordenar_tabla_productos('Código'))
                
            # Configurar ancho de columnas
            self.tabla_productos.column('ID', width=50)
            self.tabla_productos.column('Nombre', width=200)
            self.tabla_productos.column('Precio', width=80)
            self.tabla_productos.column('Costo', width=80)
            if self.stock_habilitado():
                self.tabla_productos.column('Stock', width=60)
            self.tabla_productos.column('Sugerido', width=80)
            self.tabla_productos.column('Categoría', width=100)
            self.tabla_productos.column('Código', width=120)
            
            self.tabla_productos.pack(side='left', fill='both', expand=True)
            self.tabla_productos.bind('<Double-Button-1>', self.editar_producto)
        
        # Actualizar tabla y lista de productos
        self.actualizar_tabla_productos()
        if hasattr(self, 'actualizar_lista_productos'):
            self.actualizar_lista_productos()
        
        mensaje = "Stock habilitado" if self.stock_var.get() else "Stock deshabilitado"
        messagebox.showinfo("Configuración", f"{mensaje} correctamente")
    
    def actualizar_visibilidad_stock(self):
        """Actualiza la visibilidad de los campos relacionados con stock"""
        if self.stock_habilitado():
            self.stock_label.pack(pady=2)
            self.prod_stock.pack(pady=2)
        else:
            self.stock_label.pack_forget()
            self.prod_stock.pack_forget()
    
    def guardar_ganancia_deseada_global(self):
        """Guarda la configuración global de ganancia deseada"""
        try:
            ganancia_str = self.spinbox_ganancia_deseada.get().strip()
            ganancia = float(ganancia_str) if ganancia_str else 30.0
            
            # Validar que sea un porcentaje válido
            if ganancia < 0 or ganancia >= 100:
                messagebox.showerror("Error", "La ganancia deseada debe ser un valor entre 0 y 99")
                self.spinbox_ganancia_deseada.delete(0, tk.END)
                self.spinbox_ganancia_deseada.insert(0, '30')
                return
            
            # Guardar en configuración
            self.set_configuracion('ganancia_deseada_default', str(ganancia))
            messagebox.showinfo("Éxito", f"Ganancia deseada global establecida a {ganancia}%")
            
            # Recalcular precios sugeridos para todos los productos sin override
            self.cursor.execute('''
                UPDATE productos 
                SET precio_sugerido = CASE 
                    WHEN costo > 0 THEN costo / (1 - ? / 100.0) 
                    ELSE 0 
                END
                WHERE ganancia_deseada IS NULL
            ''', (ganancia,))
            self.conn.commit()
            
            # Actualizar tabla para mostrar los nuevos precios sugeridos
            self.actualizar_tabla_productos()
        except ValueError:
            messagebox.showerror("Error", "La ganancia deseada debe ser un número válido")
    
    def crear_pestaña_reportes(self):
        """Crea la pestaña de reportes"""
        frame_reportes = tk.Frame(self.notebook, bg='#FAF2E3')
        self.notebook.add(frame_reportes, text='📊 Reportes')
        
        # Frame superior - Resumen
        frame_resumen = tk.Frame(frame_reportes, bg='#FAF2E3')
        frame_resumen.pack(fill='x', padx=10, pady=10)
        
        tk.Label(
            frame_resumen,
            text="Resumen de Ventas",
            font=('Arial', 16, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=10)
        
        # Tarjetas de resumen
        frame_tarjetas = tk.Frame(frame_resumen, bg='#FAF2E3')
        frame_tarjetas.pack(fill='x', pady=10)
        
        # Calcular estadísticas
        # self.actualizar_estadisticas()
        
        # Hoy
        self.card_hoy = tk.Frame(frame_tarjetas, bg='#dbeafe', relief='raised', bd=2)
        self.card_hoy.pack(side='left', fill='both', expand=True, padx=5)
        
        tk.Label(self.card_hoy, text="Ventas Hoy", font=('Arial', 12, 'bold'), bg='#dbeafe').pack(pady=5)
        self.label_hoy_total = tk.Label(self.card_hoy, text="$0", font=('Arial', 18, 'bold'), bg='#dbeafe', fg='#2563eb')
        self.label_hoy_total.pack()
        self.label_hoy_ventas = tk.Label(self.card_hoy, text="0 ventas", font=('Arial', 10), bg='#dbeafe')
        self.label_hoy_ventas.pack()
        self.label_hoy_ganancia = tk.Label(self.card_hoy, text="Ganancia: $0", font=('Arial', 11, 'bold'), bg='#dbeafe', fg='#16a34a')
        self.label_hoy_ganancia.pack(pady=5)
        
        # Mes
        self.card_mes = tk.Frame(frame_tarjetas, bg='#e0e7ff', relief='raised', bd=2)
        self.card_mes.pack(side='left', fill='both', expand=True, padx=5)
        
        tk.Label(self.card_mes, text="Ventas Este Mes", font=('Arial', 12, 'bold'), bg='#e0e7ff').pack(pady=5)
        self.label_mes_total = tk.Label(self.card_mes, text="$0", font=('Arial', 18, 'bold'), bg='#e0e7ff', fg='#4f46e5')
        self.label_mes_total.pack()
        self.label_mes_ventas = tk.Label(self.card_mes, text="0 ventas", font=('Arial', 10), bg='#e0e7ff')
        self.label_mes_ventas.pack()
        self.label_mes_ganancia = tk.Label(self.card_mes, text="Ganancia: $0", font=('Arial', 11, 'bold'), bg='#e0e7ff', fg='#16a34a')
        self.label_mes_ganancia.pack(pady=5)
        
        # Año
        self.card_anio = tk.Frame(frame_tarjetas, bg='#fce7f3', relief='raised', bd=2)
        self.card_anio.pack(side='left', fill='both', expand=True, padx=5)
        
        tk.Label(self.card_anio, text="Ventas Este Año", font=('Arial', 12, 'bold'), bg='#fce7f3').pack(pady=5)
        self.label_anio_total = tk.Label(self.card_anio, text="$0", font=('Arial', 18, 'bold'), bg='#fce7f3', fg='#ec4899')
        self.label_anio_total.pack()
        self.label_anio_ventas = tk.Label(self.card_anio, text="0 ventas", font=('Arial', 10), bg='#fce7f3')
        self.label_anio_ventas.pack()
        self.label_anio_ganancia = tk.Label(self.card_anio, text="Ganancia: $0", font=('Arial', 11, 'bold'), bg='#fce7f3', fg='#16a34a')
        self.label_anio_ganancia.pack(pady=5)
        
        # Frame tabla de ventas
        tk.Label(
            frame_reportes,
            text="Historial de Ventas",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=10)
        
        frame_tabla_ventas = tk.Frame(frame_reportes, bg='#FAF2E3')
        frame_tabla_ventas.pack(fill='both', expand=True, padx=10, pady=10)
        
        scrollbar_ventas = tk.Scrollbar(frame_tabla_ventas)
        scrollbar_ventas.pack(side='right', fill='y')
        
        self.tabla_ventas = ttk.Treeview(
            frame_tabla_ventas,
            columns=('ID', 'Fecha', 'Usuario', 'Método', 'Total', 'Ganancia', 'Turno'),
            show='headings',
            yscrollcommand=scrollbar_ventas.set
        )
        
        self.tabla_ventas.heading('ID', text='ID')
        self.tabla_ventas.heading('Fecha', text='Fecha')
        self.tabla_ventas.heading('Usuario', text='Usuario')
        self.tabla_ventas.heading('Método', text='Método Pago')
        self.tabla_ventas.heading('Total', text='Total')
        self.tabla_ventas.heading('Ganancia', text='Ganancia')
        self.tabla_ventas.heading('Turno', text='Turno')

        self.tabla_ventas.column('ID', width=50)
        self.tabla_ventas.column('Fecha', width=150)
        self.tabla_ventas.column('Usuario', width=150)
        self.tabla_ventas.column('Método', width=120)
        self.tabla_ventas.column('Total', width=100)
        self.tabla_ventas.column('Ganancia', width=100)
        self.tabla_ventas.column('Turno', width=80)
        
        self.tabla_ventas.pack(side='left', fill='both', expand=True)
        scrollbar_ventas.config(command=self.tabla_ventas.yview)
        
        # Botones de exportación
        frame_exportar = tk.Frame(frame_reportes, bg='#FAF2E3')
        frame_exportar.pack(fill='x', padx=10, pady=10)
        
        tk.Button(
            frame_exportar,
            text="Exportar Todo",
            font=('Arial', 11, 'bold'),
            bg='#16a34a',
            fg='white',
            command=self.exportar_todo_excel,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Button(
            frame_exportar,
            text="Actualizar Reportes",
            font=('Arial', 11),
            bg='#2563eb',
            fg='white',
            command=self.actualizar_estadisticas,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Label(
            frame_exportar,
            text="Fecha (YYYY-MM-DD):",
            font=('Arial', 10),
            bg='#FAF2E3'
        ).pack(side='left', padx=5)

        self.entry_fecha_reporte = tk.Entry(frame_exportar, font=('Arial', 10), width=12)
        self.entry_fecha_reporte.pack(side='left', padx=5)

        tk.Button(
            frame_exportar,
            text="Exportar Ventas X Día",
            font=('Arial', 10, 'bold'),
            bg='#0ea5e9',
            fg='white',
            command=self.exportar_ventas_dia_excel,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Button(
            frame_exportar,
            text="Exportar Ventas X Mes",
            font=('Arial', 10, 'bold'),
            bg='#16a34a',
            fg='white',
            command=self.exportar_ventas_mes_excel,
            cursor='hand2'
        ).pack(side='left', padx=5)
        
        tk.Label(
            frame_exportar,
            text="Turno:",
            font=('Arial', 10),
            bg='#FAF2E3'
        ).pack(side='left', padx=5)
        self.combo_turno_reporte = ttk.Combobox(frame_exportar, values=['', 'MAÑANA', 'TARDE', 'NOCHE'], width=10, state='readonly')
        self.combo_turno_reporte.pack(side='left', padx=5)
        self.combo_turno_reporte.set('')
        tk.Button(
            frame_exportar,
            text="Filtrar",
            font=('Arial', 10),
            bg='#2563eb',
            fg='white',
            command=self.actualizar_tabla_ventas,
            cursor='hand2'
        ).pack(side='left', padx=10)
        
        # Separador visual antes del botón peligroso
        # tk.Label(
        #     frame_exportar,
        #     text=" | ",
        #     font=('Arial', 12, 'bold'),
        #     bg='#FAF2E3',
        #     fg='gray'
        # ).pack(side='left', padx=10)
        
        # Botón para eliminar reportes (peligroso)
        tk.Button(
            frame_exportar,
            text="ELIMINAR REPORTES",
            font=('Arial', 10, 'bold'),
            bg='#dc2626',
            fg='white',
            command=self.confirmar_eliminar_reportes,
            cursor='hand2',
            relief='raised',
            bd=3,
            width=18
        ).pack(side='right', padx=10)
                
        # Calcular estadísticas
        self.actualizar_estadisticas()
        # Cargar ventas
        self.actualizar_tabla_ventas()
    
    def crear_pestaña_usuarios(self):
        """Crea la pestaña de gestión de usuarios (solo admin)"""
        frame_usuarios = tk.Frame(self.notebook, bg='#FAF2E3')
        self.notebook.add(frame_usuarios, text='👥 Usuarios')
        
        # Frame izquierdo - Formulario
        frame_form_user = tk.Frame(frame_usuarios, bg='#FAF2E3', width=350)
        frame_form_user.pack(side='left', fill='y', padx=10, pady=10)
        frame_form_user.pack_propagate(False)
        
        tk.Label(
            frame_form_user,
            text="Agregar Usuario",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=10)
        
        tk.Label(frame_form_user, text="Nombre de Usuario:", bg='#FAF2E3').pack(pady=5)
        self.user_nombre = tk.Entry(frame_form_user, font=('Arial', 11), width=30)
        self.user_nombre.pack(pady=5)
        
        tk.Label(frame_form_user, text="Contraseña:", bg='#FAF2E3').pack(pady=5)
        self.user_password = tk.Entry(frame_form_user, font=('Arial', 11), width=30, show='*')
        self.user_password.pack(pady=5)
        
        tk.Label(frame_form_user, text="Rol:", bg='#FAF2E3').pack(pady=5)
        self.user_rol = ttk.Combobox(
            frame_form_user,
            font=('Arial', 11),
            width=28,
            values=['empleado', 'admin'],
            state='readonly'
        )
        self.user_rol.set('empleado')
        self.user_rol.pack(pady=5)
        
        tk.Button(
            frame_form_user,
            text="Agregar Usuario",
            font=('Arial', 11, 'bold'),
            bg='#2563eb',
            fg='white',
            command=self.agregar_usuario,
            cursor='hand2'
        ).pack(pady=20)
        
        # Frame derecho - Lista de usuarios
        frame_lista_users = tk.Frame(frame_usuarios, bg='#FAF2E3')
        frame_lista_users.pack(side='right', fill='both', expand=True, padx=10, pady=10)
        
        tk.Label(
            frame_lista_users,
            text="Lista de Usuarios",
            font=('Arial', 14, 'bold'),
            bg='#FAF2E3'
        ).pack(pady=10)
        
        frame_tabla_users = tk.Frame(frame_lista_users, bg='#FAF2E3')
        frame_tabla_users.pack(fill='both', expand=True)
        
        scrollbar_users = tk.Scrollbar(frame_tabla_users)
        scrollbar_users.pack(side='right', fill='y')
        
        self.tabla_usuarios = ttk.Treeview(
            frame_tabla_users,
            columns=('ID', 'Nombre', 'Contraseña', 'Rol'),
            show='headings',
            yscrollcommand=scrollbar_users.set
        )
        
        self.tabla_usuarios.heading('ID', text='ID')
        self.tabla_usuarios.heading('Nombre', text='Nombre')
        self.tabla_usuarios.heading('Contraseña', text='Contraseña')
        self.tabla_usuarios.heading('Rol', text='Rol')
        
        self.tabla_usuarios.column('ID', width=80)
        self.tabla_usuarios.column('Nombre', width=250)
        self.tabla_usuarios.column('Contraseña', width=200)
        self.tabla_usuarios.column('Rol', width=150)
        
        self.tabla_usuarios.pack(side='left', fill='both', expand=True)
        scrollbar_users.config(command=self.tabla_usuarios.yview)
        
        tk.Button(
            frame_lista_users,
            text="Eliminar Usuario Seleccionado",
            font=('Arial', 10),
            bg='#dc2626',
            fg='white',
            command=self.eliminar_usuario,
            cursor='hand2'
        ).pack(pady=10)
        tk.Button(
            frame_lista_users,
            text="Editar Usuario Seleccionado",
            font=('Arial', 10),
            bg='#ea580c',
            fg='white',
            command=self.editar_usuario,
            cursor='hand2'
        ).pack(pady=5) 
        self.actualizar_tabla_usuarios()
    
    # ===== MÉTODOS DE VENTA =====
    
    def actualizar_lista_productos(self):
        """Actualiza la lista de productos en el punto de venta"""
        busqueda = self.entry_buscar.get().lower()
        self.lista_productos.delete(0, tk.END)
        
        if busqueda:
            self.cursor.execute('''
                SELECT * FROM productos 
                WHERE LOWER(nombre) LIKE ? OR LOWER(categoria) LIKE ?
                ORDER BY nombre
            ''', (f'%{busqueda}%', f'%{busqueda}%'))
        else:
            self.cursor.execute('SELECT * FROM productos ORDER BY nombre')
        
        productos = self.cursor.fetchall()
        for producto in productos:
            if self.stock_habilitado():
                # Mostrar con información de stock
                texto = f"{producto[1]} - ${producto[2]} - Stock: {producto[4]} - {producto[5]}"
                # Determinar color según prioridades
                if producto[2] == 0 or producto[3] == 0:  # precio o costo en cero
                    color = '#fef3c7'  # Amarillo (prioridad alta)
                elif producto[4] <= 5:  # stock bajo
                    color = '#fee2e2'  # Rojo claro
                else:
                    color = 'white'
            else:
                # Mostrar sin información de stock
                texto = f"{producto[1]} - ${producto[2]} - {producto[5]}"
                # Color amarillo si precio o costo están en cero
                color = '#fef3c7' if producto[2] == 0 or producto[3] == 0 else 'white'
            
            self.lista_productos.insert(tk.END, texto)
            self.lista_productos.itemconfig(tk.END, {'bg': color})
    
    def buscar_por_barcode(self, event):
        """Busca y agrega producto por código de barras"""
        codigo = self.entry_barcode.get()
        if not codigo:
            return
        
        self.cursor.execute('SELECT * FROM productos WHERE codigo_barras = ?', (codigo,))
        producto = self.cursor.fetchone()
        
        if producto:
            # Solo validar stock si está habilitado
            if self.stock_habilitado() and producto[4] <= 0:
                messagebox.showwarning("Sin Stock", "No hay stock disponible de este producto")
                return
            
            # Verificar si ya está en el carrito
            for item in self.carrito:
                if item['id'] == producto[0]:
                    # Solo validar límite de stock si está habilitado
                    if self.stock_habilitado() and item['cantidad'] >= producto[4]:
                        messagebox.showwarning("Stock Insuficiente", "No hay más stock disponible")
                        return
                    item['cantidad'] += 1
                    self.actualizar_carrito_display()
                    self.entry_barcode.delete(0, tk.END)
                    return
            
            # Agregar nuevo item
            self.carrito.append({
                'id': producto[0],
                'nombre': producto[1],
                'precio': producto[2],
                'costo': producto[3],
                'cantidad': 1,
                'stock_disponible': producto[4]
            })
            self.actualizar_carrito_display()
            self.entry_barcode.delete(0, tk.END)
        else:
            messagebox.showerror("No encontrado", "Producto no encontrado con ese código de barras")
    
    def agregar_al_carrito(self):
        """Agrega el producto seleccionado al carrito"""
        seleccion = self.lista_productos.curselection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un producto")
            return
        
        # Obtener el producto de la base de datos
        busqueda = self.entry_buscar.get().lower()
        if busqueda:
            self.cursor.execute('''
                SELECT * FROM productos 
                WHERE LOWER(nombre) LIKE ? OR LOWER(categoria) LIKE ?
                ORDER BY nombre
            ''', (f'%{busqueda}%', f'%{busqueda}%'))
        else:
            self.cursor.execute('SELECT * FROM productos ORDER BY nombre')
        
        productos = self.cursor.fetchall()
        producto = productos[seleccion[0]]
        
        # Solo validar stock si está habilitado
        if self.stock_habilitado() and producto[4] <= 0:
            messagebox.showwarning("Sin Stock", "No hay stock disponible de este producto")
            return
        
        # Verificar si ya está en el carrito
        for item in self.carrito:
            if item['id'] == producto[0]:
                # Solo validar límite de stock si está habilitado
                if self.stock_habilitado() and item['cantidad'] >= producto[4]:
                    messagebox.showwarning("Stock Insuficiente", "No hay más stock disponible")
                    return
                item['cantidad'] += 1
                self.actualizar_carrito_display()
                return
        
        # Agregar nuevo item
        self.carrito.append({
            'id': producto[0],
            'nombre': producto[1],
            'precio': producto[2],
            'costo': producto[3],
            'cantidad': 1,
            'stock_disponible': producto[4]
        })
        self.actualizar_carrito_display()
    
    def actualizar_carrito_display(self):
        """Actualiza la visualización del carrito"""
        self.lista_carrito.delete(0, tk.END)
        total = 0
        
        for item in self.carrito:
            subtotal = item['precio'] * item['cantidad']
            total += subtotal
            texto = f"{item['nombre']} x{item['cantidad']} - ${subtotal}"
            self.lista_carrito.insert(tk.END, texto)
        
        self.label_total.config(text=f"TOTAL: ${total}")
    
    def eliminar_del_carrito(self):
        """Elimina el item seleccionado del carrito"""
        seleccion = self.lista_carrito.curselection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un item del carrito")
            return
        print(self.carrito[seleccion[0]]['id'])
        #logging.info(seleccion.id)
        # messagebox.showwarning(seleccion[0])
        del self.carrito[seleccion[0]]
        self.actualizar_carrito_display()

    def eliminar_uno_del_carrito(self):
        """Elimina el item seleccionado del carrito"""
        seleccion = self.lista_carrito.curselection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un item del carrito")
            return
        index = seleccion[0]
        producto = self.carrito[index]

        # Restar una unidad
        producto["cantidad"] -= 1

        # Si llega a 0, eliminar completamente
        if producto["cantidad"] <= 0:
            del self.carrito[index]

        # Actualizar la interfaz
        self.actualizar_carrito_display()
        self.lista_carrito.selection_set(index)

    def vaciar_carrito(self):
        """Vacía completamente el carrito"""
        if self.carrito and messagebox.askyesno("Confirmar", "¿Vaciar el carrito?"):
            self.carrito = []
            self.actualizar_carrito_display()
    
    def restar_stock_interno(self):
        """Resta stock usando los productos del carrito sin registrar una venta."""
        if not self.carrito:
            messagebox.showwarning("Carrito vacío", "No hay productos en el carrito para restar stock.")
            return

        # Solo permitir restar stock si está habilitado
        if not self.stock_habilitado():
            messagebox.showinfo("Stock deshabilitado", "El control de stock está deshabilitado. No se puede restar stock.")
            return

        if not messagebox.askyesno("Confirmar", "¿Deseas restar el stock de los productos del carrito sin registrar una venta?"):
            return

        for item in self.carrito:
            self.cursor.execute('''
                UPDATE productos SET stock = stock - ? WHERE id = ?
            ''', (item['cantidad'], item['id']))

        self.conn.commit()

        self.carrito = []
        self.actualizar_carrito_display()
        if hasattr(self, 'actualizar_lista_productos'):
            self.actualizar_lista_productos()

        messagebox.showinfo("Stock actualizado", "El stock fue actualizado correctamente (sin registrar venta).")
    def cobrar_metodo(self, metodo):
        """Wrapper para procesar cobros: confirma para no-efectivo, ventana especial para efectivo."""
        # Calcular total actual del carrito
        if not self.carrito:
            messagebox.showwarning("Carrito Vacío", "El carrito está vacío")
            return

        total = sum(item['precio'] * item['cantidad'] for item in self.carrito)

        if metodo and metodo.lower() == 'efectivo':
            # Mostrar diálogo para ingresar el pago y calcular vuelto
            self.mostrar_dialogo_efectivo(total, metodo)
        else:
            # Confirmar para otros métodos con diálogo que acepta F1
            self.mostrar_dialogo_confirmacion(total, metodo)

    def mostrar_dialogo_efectivo(self, total, metodo='Efectivo'):
        """Muestra una ventana modal para ingresar pago en efectivo y calcular vuelto en vivo."""
        # Deshabilitar atajos globales mientras el diálogo esté abierto
        self.shortcuts_enabled = False

        dlg = tk.Toplevel(self.root)
        dlg.title("Cobro en Efectivo")
        dlg.geometry("360x200")
        dlg.configure(bg='#FAF2E3')
        dlg.transient(self.root)
        dlg.grab_set()

        # Asegurar que al cerrar con la X se re-habiliten los atajos globales
        def _on_close_efectivo():
            self.shortcuts_enabled = True
            try:
                dlg.grab_release()
            except:
                pass
            dlg.destroy()
        dlg.protocol("WM_DELETE_WINDOW", _on_close_efectivo)
        # Asegurar que el diálogo reciba el foco para capturar F1/F2
        try:
            dlg.focus_set()
            dlg.focus_force()
        except Exception:
            pass

        tk.Label(dlg, text=f"Total a pagar: ${total:.2f}", font=('Arial', 12, 'bold'), bg='#FAF2E3').pack(pady=(12,6))

        frame_pago = tk.Frame(dlg, bg='#FAF2E3')
        frame_pago.pack(pady=6, padx=10, fill='x')

        tk.Label(frame_pago, text="Pago recibido:", font=('Arial', 10), bg='#FAF2E3').pack(anchor='w')
        entry_pago = tk.Entry(frame_pago, font=('Arial', 12))
        entry_pago.pack(fill='x', pady=4)

        label_vuelto = tk.Label(dlg, text="Vuelto: $0.00", font=('Arial', 12, 'bold'), bg='#FAF2E3')
        label_vuelto.pack(pady=6)

        def actualizar_vuelto(event=None):
            val = entry_pago.get().strip()
            # Si el campo está vacío, mostrar vuelto como '-' (opcional para el empleado)
            if val == '':
                label_vuelto.config(text="Vuelto: -")
                return
            try:
                pago = float(val)
            except Exception:
                label_vuelto.config(text="Vuelto: -")
                return
            vuelto = pago - total
            label_vuelto.config(text=f"Vuelto: ${vuelto:.2f}")

        entry_pago.bind('<KeyRelease>', actualizar_vuelto)

        def confirmar():
            # Re-habilitar atajos antes de cerrar
            self.shortcuts_enabled = True
            val = entry_pago.get().strip()
            # Campo vacío -> proceder igual (el cálculo de vuelto es opcional)
            if val == '':
                try:
                    dlg.grab_release()
                except:
                    pass
                dlg.destroy()
                self.finalizar_venta(metodo)
                return

            # Si llenaron algo, validar que sea numérico
            try:
                pago = float(val)
            except Exception:
                messagebox.showerror("Error", "Ingrese un monto válido para el pago")
                return

            # Si el pago es menor al total, preguntar si registrar igual
            #if pago < total:
            #    if not messagebox.askyesno("Pago insuficiente", "El pago ingresado es menor al total. ¿Desea registrar la venta igual?", icon='warning'):
            #        return

            dlg.grab_release()
            dlg.destroy()
            # Proceder con la venta
            self.finalizar_venta(metodo)

        frame_bot = tk.Frame(dlg, bg='#FAF2E3')
        frame_bot.pack(pady=8, fill='x', padx=10)

        tk.Button(frame_bot, text="Confirmar Cobro", bg='#16a34a', fg='white', font=('Arial', 10, 'bold'), command=confirmar).pack(side='left', expand=True, fill='x', padx=5)
        def _cancelar_efectivo():
            self.shortcuts_enabled = True
            try:
                dlg.grab_release()
            except:
                pass
            dlg.destroy()
        tk.Button(frame_bot, text="Cancelar", bg='#ef4444', fg='white', font=('Arial', 10), command=_cancelar_efectivo).pack(side='left', expand=True, fill='x', padx=5)

        entry_pago.focus_set()
        # Inicializar vuelto
        actualizar_vuelto()
        # Permitir confirmar con F1 dentro del diálogo
        try:
            # Bind F1 to confirmar and F2 to cancelar inside the dialog. Return 'break' to stop propagation to root bindings.
            dlg.bind('<F1>', lambda e: (confirmar(), 'break'))
            dlg.bind('<F2>', lambda e: (_cancelar_efectivo(), 'break'))
        except Exception:
            pass

    def mostrar_dialogo_confirmacion(self, total, metodo=''):
        """Muestra un diálogo modal de confirmación con la posibilidad de confirmar con F1."""
        # Deshabilitar atajos globales mientras el diálogo esté abierto
        self.shortcuts_enabled = False

        dlg = tk.Toplevel(self.root)
        dlg.title("Confirmar Cobro")
        dlg.geometry("360x150")
        dlg.configure(bg='#FAF2E3')
        dlg.transient(self.root)
        dlg.grab_set()

        # Asegurar que al cerrar con la X se re-habiliten los atajos globales
        def _on_close_confirm():
            self.shortcuts_enabled = True
            try:
                dlg.grab_release()
            except:
                pass
            dlg.destroy()
        dlg.protocol("WM_DELETE_WINDOW", _on_close_confirm)
        # Asegurar que el diálogo reciba el foco para capturar F1/F2
        try:
            dlg.focus_set()
            dlg.focus_force()
        except Exception:
            pass

        tk.Label(dlg, text=f"Confirmar cobro por {metodo}", font=('Arial', 12, 'bold'), bg='#FAF2E3').pack(pady=(12,6))
        tk.Label(dlg, text=f"Total: ${total:.2f}", font=('Arial', 11), bg='#FAF2E3').pack(pady=(0,8))

        def confirmar():
            self.shortcuts_enabled = True
            try:
                dlg.grab_release()
            except:
                pass
            dlg.destroy()
            self.finalizar_venta(metodo)

        frame_bot = tk.Frame(dlg, bg='#FAF2E3')
        frame_bot.pack(pady=8, fill='x', padx=10)

        tk.Button(frame_bot, text="F1  Confirmar", bg='#16a34a', fg='white', font=('Arial', 10, 'bold'), command=confirmar).pack(side='left', expand=True, fill='x', padx=5)
        def _cancelar_confirm():
            self.shortcuts_enabled = True
            try:
                dlg.grab_release()
            except:
                pass
            dlg.destroy()
        tk.Button(frame_bot, text="F2  Cancelar", bg='#ef4444', fg='white', font=('Arial', 10), command=_cancelar_confirm).pack(side='left', expand=True, fill='x', padx=5)

        # Bind F1 to confirmar and F2 to cancelar inside the dialog. Return 'break' to stop propagation to root bindings.
        try:
            dlg.bind('<F1>', lambda e: (confirmar(), 'break'))
            dlg.bind('<F2>', lambda e: (_cancelar_confirm(), 'break'))
        except Exception:
            pass
    def finalizar_venta(self, metodo_pago):
        """Finaliza la venta y la registra"""
        if not self.carrito:
            messagebox.showwarning("Carrito Vacío", "El carrito está vacío")
            return
        
        # Calcular totales
        total = sum(item['precio'] * item['cantidad'] for item in self.carrito)
        costo_total = sum(item['costo'] * item['cantidad'] for item in self.carrito)
        
        # Registrar venta con turno
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        turno = self.turno_actual if hasattr(self, 'turno_actual') and self.turno_actual else 'MAÑANA'
        self.cursor.execute('''
            INSERT INTO ventas (fecha, usuario, metodo_pago, total, costo_total, turno)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (fecha, self.usuario_actual['nombre'], metodo_pago, total, costo_total, turno))

        venta_id = self.cursor.lastrowid
        
        # Registrar items de la venta
        for item in self.carrito:
            self.cursor.execute('''
                INSERT INTO items_venta (venta_id, producto_nombre, cantidad, precio_unitario, costo_unitario)
                VALUES (?, ?, ?, ?, ?)
            ''', (venta_id, item['nombre'], item['cantidad'], item['precio'], item['costo']))
            
            # Solo actualizar stock si está habilitado
            if self.stock_habilitado():
                self.cursor.execute('''
                    UPDATE productos SET stock = stock - ? WHERE id = ?
                ''', (item['cantidad'], item['id']))
        
        self.conn.commit()
        
        # Generar ticket
        if messagebox.askyesno("Ticket", "¿Deseas generar el ticket de venta?"):
            self.generar_ticket(venta_id, metodo_pago, total)
        
        # Limpiar carrito
        self.carrito = []
        self.actualizar_carrito_display()
        if hasattr(self, 'actualizar_lista_productos'):
            self.actualizar_lista_productos()
        
        messagebox.showinfo("Éxito", f"Venta registrada exitosamente\nTotal: ${total}")
    
    def generar_ticket(self, venta_id, metodo_pago, total):
        """Genera un ticket PDF de la venta"""
        try:
            # Crear carpeta 'tickets' si no existe
            carpeta_tickets = "tickets"
            if not os.path.exists(carpeta_tickets):
                os.makedirs(carpeta_tickets)

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial', 'B', 16)
            
            # Encabezado
            pdf.cell(0, 10, 'KIOSCO - TICKET DE VENTA', 0, 1, 'C')
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 5, f'Ticket N: {venta_id}', 0, 1, 'C')
            pdf.cell(0, 5, f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'C')
            pdf.cell(0, 5, f'Vendedor: {self.usuario_actual["nombre"]}', 0, 1, 'C')
            pdf.ln(5)
            
            # Línea separadora
            pdf.cell(0, 0, '', 'T', 1)
            pdf.ln(3)
            
            # Items
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(100, 5, 'Producto', 0, 0)
            pdf.cell(30, 5, 'Cant.', 0, 0, 'C')
            pdf.cell(40, 5, 'Precio', 0, 1, 'R')
            
            pdf.set_font('Arial', '', 9)
            self.cursor.execute('''
                SELECT producto_nombre, cantidad, precio_unitario
                FROM items_venta WHERE venta_id = ?
            ''', (venta_id,))
            
            items = self.cursor.fetchall()
            for item in items:
                pdf.cell(100, 5, item[0][:30], 0, 0)
                pdf.cell(30, 5, str(item[1]), 0, 0, 'C')
                pdf.cell(40, 5, f'${item[2] * item[1]:.2f}', 0, 1, 'R')
            
            pdf.ln(3)
            pdf.cell(0, 0, '', 'T', 1)
            pdf.ln(3)
            
            # Total
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(130, 8, 'TOTAL:', 0, 0, 'R')
            pdf.cell(40, 8, f'${total:.2f}', 0, 1, 'R')
            
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 5, f'Metodo de pago: {metodo_pago}', 0, 1, 'C')
            
            pdf.ln(10)
            pdf.set_font('Arial', 'I', 8)
            pdf.cell(0, 5, 'Gracias por su compra!', 0, 1, 'C')
            
            # Guardar en subcarpeta 'tickets'
            filename = os.path.join(carpeta_tickets, f'ticket_{venta_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf')
            pdf.output(filename)
            messagebox.showinfo("Ticket Generado", f"Ticket guardado como: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar ticket: {str(e)}")
    
    # ===== MÉTODOS DE PRODUCTOS =====
    
    def guardar_producto(self):
        """Guarda o actualiza un producto"""
        nombre = self.prod_nombre.get()
        precio = self.prod_precio.get()
        costo = self.prod_costo.get()
        categoria = self.prod_categoria.get()
        codigo_barras = self.prod_barcode.get()
        
        # Validar que al menos tenga nombre
        if not nombre or nombre.strip() == '':
            messagebox.showwarning("Campo Vacío", "El nombre del producto es obligatorio")
            return
        
        # Manejar stock según configuración
        if self.stock_habilitado():
            stock = self.prod_stock.get()
            if not stock:
                messagebox.showwarning("Campo Stock", "El campo stock es obligatorio cuando está habilitado")
                return
            try:
                stock = int(stock)
            except ValueError:
                messagebox.showerror("Error", "El stock debe ser un número entero válido")
                return
        else:
            stock = 0  # Valor por defecto cuando stock está deshabilitado
        
        # Convertir precio y costo, permitiendo valores vacíos (se convertirán a 0)
        try:
            precio = float(precio) if precio and precio.strip() != '' else 0.0
            costo = float(costo) if costo and costo.strip() != '' else 0.0
        except ValueError:
            messagebox.showerror("Error", "Precio y costo deben ser números válidos (o estar vacíos)")
            return
        
        # Obtener ganancia_deseada si está disponible
        ganancia_deseada = None
        if hasattr(self, 'prod_ganancia_deseada'):
            ganancia_str = self.prod_ganancia_deseada.get().strip()
            if ganancia_str:
                try:
                    ganancia_deseada = float(ganancia_str)
                except ValueError:
                    messagebox.showerror("Error", "La ganancia deseada debe ser un número válido o estar vacía")
                    return
        
        # Advertir si precio o costo están en cero
        if precio == 0 or costo == 0:
            advertencia = []
            if precio == 0:
                advertencia.append("precio")
            if costo == 0:
                advertencia.append("costo")
            
            mensaje_adv = f"⚠️ El {' y '.join(advertencia)} {'está' if len(advertencia) == 1 else 'están'} en cero.\n\n"
            mensaje_adv += "El producto se guardará pero aparecerá marcado en amarillo hasta que completes todos los campos.\n\n"
            mensaje_adv += "¿Deseas continuar?"
            
            if not messagebox.askyesno("Campos Incompletos", mensaje_adv, icon='warning'):
                return
        
        # Calcular precio sugerido con la ganancia deseada (o global si no está definida)
        precio_sugerido = self.calcular_precio_sugerido(costo, ganancia_deseada)
        
        if self.producto_id:
            # Actualizar
            self.cursor.execute('''
                UPDATE productos 
                SET nombre=?, precio=?, costo=?, stock=?, categoria=?, codigo_barras=?, precio_sugerido=?, ganancia_deseada=?
                WHERE id=?
            ''', (nombre, precio, costo, stock, categoria or 'Otros', codigo_barras, precio_sugerido, ganancia_deseada, self.producto_id))
            messagebox.showinfo("Éxito", "Producto actualizado correctamente")
        else:
            # Insertar
            self.cursor.execute('''
                INSERT INTO productos (nombre, precio, costo, stock, categoria, codigo_barras, precio_sugerido, ganancia_deseada)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (nombre, precio, costo, stock, categoria or 'Otros', codigo_barras, precio_sugerido, ganancia_deseada))
            # Reorganizar IDs después de agregar nuevo producto (sin commit automático)
            self.reorganizar_ids_productos(auto_commit=False)
            self.conn.commit()
            messagebox.showinfo("Éxito", "Producto agregado correctamente")
        
        if self.producto_id:  # Solo commit si es actualización (inserción ya hizo commit)
            self.conn.commit()
        
        self.limpiar_formulario_producto()
        self.actualizar_tabla_productos()
        if hasattr(self, 'actualizar_lista_productos'):
            self.actualizar_lista_productos()
    
    def limpiar_formulario_producto(self):
        """Limpia el formulario de productos"""
        self.producto_id = None
        self.prod_nombre.delete(0, tk.END)
        self.prod_precio.delete(0, tk.END)
        self.prod_costo.delete(0, tk.END)
        self.prod_stock.delete(0, tk.END)
        self.prod_categoria.set('')
        self.prod_barcode.delete(0, tk.END)
        if hasattr(self, 'prod_ganancia_deseada'):
            self.prod_ganancia_deseada.delete(0, tk.END)
    
    def actualizar_tabla_productos(self):
        """Actualiza la tabla de productos"""
        for item in self.tabla_productos.get_children():
            self.tabla_productos.delete(item)
        # Soporte para filtro de búsqueda si el campo existe
        if hasattr(self, 'entry_buscar_productos'):
            filtro = self.entry_buscar_productos.get().lower().strip()
            if filtro:
                # Buscar por nombre, categoría o código de barras
                self.cursor.execute('''
                    SELECT * FROM productos
                    WHERE LOWER(nombre) LIKE ? OR LOWER(categoria) LIKE ? OR LOWER(COALESCE(codigo_barras, '')) LIKE ?
                    ORDER BY nombre
                ''', (f'%{filtro}%', f'%{filtro}%', f'%{filtro}%'))
            else:
                self.cursor.execute('SELECT * FROM productos ORDER BY nombre')
        else:
            self.cursor.execute('SELECT * FROM productos ORDER BY nombre')

        productos = self.cursor.fetchall()
        
        for producto in productos:
            # Limpiar código de barras si tiene .0 al final
            producto_lista = list(producto)
            if producto_lista[6] and str(producto_lista[6]).endswith('.0'):
                producto_lista[6] = str(producto_lista[6])[:-2]
            
            # Recalcular precio_sugerido si es 0 o inválido
            # Estructura: id(0), nombre(1), precio(2), costo(3), stock(4), categoria(5), codigo_barras(6), precio_sugerido(7), ganancia_deseada(8)
            costo = producto_lista[3]
            ganancia_deseada = producto_lista[8] if len(producto_lista) > 8 else None
            precio_sugerido = self.calcular_precio_sugerido(costo, ganancia_deseada)
            if precio_sugerido != producto_lista[7]:
                # Actualizar en la base de datos
                self.cursor.execute('''
                    UPDATE productos SET precio_sugerido = ? WHERE id = ?
                ''', (precio_sugerido, producto_lista[0]))
                self.conn.commit()
            producto_lista[7] = precio_sugerido
            
            producto = tuple(producto_lista)
            # Determinar tags según el estado del producto
            tags = []
            
            # Tag por precio/costo incompleto (prioridad alta)
            if producto[2] == 0 or producto[3] == 0:  # precio == 0 o costo == 0
                tags.append('incompleto')
            # Tag por precio menor al sugerido (segunda prioridad)
            elif producto[2] > 0 and producto[7] > 0 and producto[2] < producto[7]:
                tags.append('precio_bajo')
            # Tag por stock bajo (solo si stock habilitado y no está incompleto ni precio bajo)
            elif self.stock_habilitado() and producto[4] <= 5:
                tags.append('bajo_stock')
            
            if self.stock_habilitado():
                # Mostrar: ID(0), Nombre(1), Precio(2), Costo(3), Stock(4), Sugerido(7), Categoría(5), Código(6)
                valores = (producto[0], producto[1], producto[2], producto[3], producto[4], producto[7], producto[5], producto[6])
            else:
                # Omitir stock: ID(0), Nombre(1), Precio(2), Costo(3), Sugerido(7), Categoría(5), Código(6)
                valores = (producto[0], producto[1], producto[2], producto[3], producto[7], producto[5], producto[6])
            
            self.tabla_productos.insert('', 'end', values=valores, tags=tuple(tags))
        
        # Configurar colores para los tags
        self.tabla_productos.tag_configure('incompleto', background='#fef3c7', foreground='#92400e')  # Amarillo
        self.tabla_productos.tag_configure('precio_bajo', background='#fee2e2', foreground='#dc2626')  # Rojo claro
        if self.stock_habilitado():
            self.tabla_productos.tag_configure('bajo_stock', background='#fee2e2', foreground='#dc2626')  # Rojo claro
    
    def editar_producto(self, event=None):
        """Carga el producto seleccionado en el formulario para editar"""
        seleccion = self.tabla_productos.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un producto")
            return
        
        item = self.tabla_productos.item(seleccion[0])
        valores = item['values']
        
        # Siempre obtener el producto completo de la BD para acceder a ganancia_deseada
        # Estructura en Treeview con stock: ID(0), Nombre(1), Precio(2), Costo(3), Stock(4), Sugerido(5), Categoría(6), Código(7)
        # Estructura en Treeview sin stock: ID(0), Nombre(1), Precio(2), Costo(3), Sugerido(4), Categoría(5), Código(6)
        producto_id = valores[0]
        self.cursor.execute('SELECT * FROM productos WHERE id = ?', (producto_id,))
        producto_completo = self.cursor.fetchone()
        
        if not producto_completo:
            messagebox.showerror("Error", "No se pudo encontrar el producto seleccionado")
            return
        
        # Estructura de BD: id(0), nombre(1), precio(2), costo(3), stock(4), categoria(5), codigo_barras(6), precio_sugerido(7), ganancia_deseada(8)
        self.producto_id = producto_completo[0]
        self.prod_nombre.delete(0, tk.END)
        self.prod_nombre.insert(0, producto_completo[1])
        self.prod_precio.delete(0, tk.END)
        self.prod_precio.insert(0, producto_completo[2])
        self.prod_costo.delete(0, tk.END)
        self.prod_costo.insert(0, producto_completo[3])
        self.prod_stock.delete(0, tk.END)
        self.prod_stock.insert(0, producto_completo[4])
        self.prod_categoria.set(producto_completo[5])
        self.prod_barcode.delete(0, tk.END)
        self.prod_barcode.insert(0, producto_completo[6] if producto_completo[6] else '')
        
        # Cargar ganancia_deseada si existe (puede ser None)
        if hasattr(self, 'prod_ganancia_deseada'):
            self.prod_ganancia_deseada.delete(0, tk.END)
            if len(producto_completo) > 8 and producto_completo[8] is not None:
                self.prod_ganancia_deseada.insert(0, str(producto_completo[8]))
    
    
    def eliminar_producto(self):
        """Elimina el producto seleccionado"""
        seleccion = self.tabla_productos.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un producto")
            return
        
        if messagebox.askyesno("Confirmar", "¿Seguro que deseas eliminar este producto?"):
            item = self.tabla_productos.item(seleccion[0])
            producto_id = item['values'][0]  # El ID siempre está en el índice 0
            
            self.cursor.execute('DELETE FROM productos WHERE id = ?', (producto_id,))
            
            # Reorganizar IDs después de eliminar producto (sin commit automático)
            self.reorganizar_ids_productos(auto_commit=False)
            self.conn.commit()
            
            self.actualizar_tabla_productos()
            if hasattr(self, 'actualizar_lista_productos'):
                self.actualizar_lista_productos()
            messagebox.showinfo("Éxito", "Producto eliminado correctamente")
    def duplicar_producto(self):
        """Duplica el producto seleccionado creando una copia con nombre único"""
        seleccion = self.tabla_productos.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un producto para duplicar")
            return

        item = self.tabla_productos.item(seleccion[0])
        valores = item['values']
        producto_id = valores[0]

        self.cursor.execute('SELECT * FROM productos WHERE id = ?', (producto_id,))
        producto = self.cursor.fetchone()
        if not producto:
            messagebox.showerror("Error", "No se pudo encontrar el producto seleccionado en la base de datos")
            return

        nombre_original = producto[1]
        nuevo_nombre = f"{nombre_original} - copia"
        base_name = nuevo_nombre
        contador = 1
        # Asegurar que el nombre sea único (campo nombre es UNIQUE)
        while True:
            self.cursor.execute('SELECT id FROM productos WHERE nombre = ?', (nuevo_nombre,))
            if not self.cursor.fetchone():
                break
            nuevo_nombre = f"{base_name} ({contador})"
            contador += 1

        precio = producto[2]
        costo = producto[3]
        stock = producto[4]
        categoria = producto[5]
        codigo_barras = producto[6]

        self.cursor.execute('''
            INSERT INTO productos (nombre, precio, costo, stock, categoria, codigo_barras)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (nuevo_nombre, precio, costo, stock, categoria, codigo_barras))

        # Mantener consistencia con el flujo de inserción existente (reorganizar IDs si existe)
        if hasattr(self, 'reorganizar_ids_productos'):
            try:
                self.reorganizar_ids_productos(auto_commit=False)
            except Exception:
                pass

        self.conn.commit()

        messagebox.showinfo("Éxito", f"Producto duplicado como: {nuevo_nombre}")
        self.actualizar_tabla_productos()
        if hasattr(self, 'actualizar_lista_productos'):
            self.actualizar_lista_productos()
    def ordenar_tabla_productos(self, col, reverse=False):
        """Ordena la tabla de productos por la columna seleccionada"""
        l = [(self.tabla_productos.set(k, col), k) for k in self.tabla_productos.get_children('')]
        try:
            l.sort(key=lambda t: float(t[0]) if col in ['ID', 'Precio', 'Costo', 'Stock'] else t[0], reverse=reverse)
        except ValueError:
            l.sort(key=lambda t: t[0], reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tabla_productos.move(k, '', index)
        # Alterna el orden para el próximo click
        self.tabla_productos.heading(col, command=lambda: self.ordenar_tabla_productos(col, not reverse))
    def exportar_productos_excel(self):
        """Exporta los productos a Excel"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"productos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if filename:
                self.cursor.execute('SELECT * FROM productos')
                productos = self.cursor.fetchall()
                
                df = pd.DataFrame(productos, columns=['ID', 'Nombre', 'Precio', 'Costo', 'Stock', 'Categoría', 'Código Barras'])
                
                # Asegurar que los códigos de barras se exporten como texto
                df['Código Barras'] = df['Código Barras'].astype(str)
                
                # Usar ExcelWriter para controlar mejor el formato
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Productos')
                    
                    # Obtener la hoja de trabajo para formatear la columna de códigos de barras
                    worksheet = writer.sheets['Productos']
                    
                    # Formatear la columna de códigos de barras como texto
                    for row in range(2, len(df) + 2):  # Empezar desde fila 2 (después del header)
                        cell = worksheet.cell(row=row, column=7)  # Columna 7 = Código Barras
                        if cell.value and str(cell.value) != 'nan':
                            cell.number_format = '@'  # Formato de texto
                            # Limpiar .0 si existe
                            value = str(cell.value)
                            if value.endswith('.0'):
                                value = value[:-2]
                            cell.value = value
                
                messagebox.showinfo("Éxito", f"Productos exportados a: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    def importar_productos(self):
        """Importa productos desde un archivo Excel o CSV"""
        file_path = filedialog.askopenfilename(
            title="Selecciona archivo de productos",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        if not file_path:
            return
    
        try:
            if file_path.endswith('.xlsx'):
                # Leer con tipos específicos para códigos de barras como texto
                df = pd.read_excel(file_path, dtype={'Código Barras': str})
            else:
                # Para CSV también especificar tipos
                df = pd.read_csv(file_path, dtype={'Código Barras': str})
    
            # Mostrar las columnas disponibles para debug
            print(f"Columnas encontradas: {list(df.columns)}")
            
            productos_importados = 0
            productos_saltados = 0
            errores_detalle = []
            
            # Verificar configuración de stock
            stock_habilitado = self.stock_habilitado()
            
            # Espera columnas: Nombre, Precio, Costo, Stock, Categoría, Código Barras
            for index, row in df.iterrows():
                try:
                    nombre = row.get('Nombre')
                    precio = row.get('Precio')
                    costo = row.get('Costo')
                    stock = row.get('Stock')
                    categoria = row.get('Categoría', 'Otros')
                    codigo_barras = row.get('Código Barras', '')
                    
                    # Validar que al menos tenga nombre
                    if pd.isnull(nombre) or str(nombre).strip() == '':
                        productos_saltados += 1
                        errores_detalle.append(f"Fila {index + 2}: Nombre es obligatorio")
                        continue
                    
                    # Manejar stock según configuración
                    if stock_habilitado:
                        if pd.isnull(stock):
                            productos_saltados += 1
                            errores_detalle.append(f"Fila {index + 2}: Stock requerido cuando está habilitado")
                            continue
                        stock = int(float(stock))
                    else:
                        stock = 0  # Valor por defecto cuando stock está deshabilitado
                    
                    # Convertir tipos de datos
                    nombre = str(nombre).strip()
                    
                    # Manejar precio y costo - permitir valores nulos/cero
                    if pd.isnull(precio) or precio == '':
                        precio = 0.0
                    else:
                        precio = float(precio)
                    
                    if pd.isnull(costo) or costo == '':
                        costo = 0.0
                    else:
                        costo = float(costo)
                    
                    categoria = str(categoria) if not pd.isnull(categoria) else 'Otros'
                    
                    # Manejar código de barras - remover .0 si existe
                    if pd.isnull(codigo_barras) or codigo_barras == '':
                        codigo_barras = ''
                    else:
                        codigo_barras = str(codigo_barras)
                        # Si termina en .0, removerlo (problema común con Excel/pandas)
                        if codigo_barras.endswith('.0'):
                            codigo_barras = codigo_barras[:-2]
                    
                    # Insertar en base de datos
                    self.cursor.execute('''
                        INSERT INTO productos (nombre, precio, costo, stock, categoria, codigo_barras)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (nombre, precio, costo, stock, categoria, codigo_barras))
                    
                    productos_importados += 1
                    
                except sqlite3.IntegrityError as e:
                    productos_saltados += 1
                    errores_detalle.append(f"Fila {index + 2}: Producto duplicado o error de integridad")
                    continue
                except ValueError as e:
                    productos_saltados += 1
                    errores_detalle.append(f"Fila {index + 2}: Error de formato de datos - {str(e)}")
                    continue
                except Exception as e:
                    productos_saltados += 1
                    errores_detalle.append(f"Fila {index + 2}: Error inesperado - {str(e)}")
                    continue
    
            self.conn.commit()
            self.actualizar_tabla_productos()
            if hasattr(self, 'actualizar_lista_productos'):
                self.actualizar_lista_productos()
            
            # Mensaje detallado de resultado
            mensaje = f"Importación completada:\n\n"
            mensaje += f"✅ Productos importados: {productos_importados}\n"
            if productos_saltados > 0:
                mensaje += f"⚠️ Productos saltados: {productos_saltados}\n\n"
                if errores_detalle:
                    mensaje += "Detalles de errores:\n"
                    # Mostrar solo los primeros 5 errores para no hacer el mensaje muy largo
                    for error in errores_detalle[:5]:
                        mensaje += f"• {error}\n"
                    if len(errores_detalle) > 5:
                        mensaje += f"... y {len(errores_detalle) - 5} errores más"
            
            if productos_importados > 0:
                # Reorganizar IDs después de la importación
                self.reorganizar_ids_productos()
                messagebox.showinfo("Importación Completada", mensaje)
            else:
                messagebox.showwarning("Sin Productos Importados", mensaje)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar archivo: {str(e)}\n\nVerifica que el archivo tenga las columnas correctas:\n• Nombre\n• Precio\n• Costo\n• Stock (si está habilitado)\n• Categoría\n• Código Barras")
    
    def reorganizar_ids_productos(self, auto_commit=True):
        """Reorganiza los IDs de productos para que sean secuenciales"""
        try:
            # Obtener todos los productos ordenados por ID actual
            self.cursor.execute('SELECT * FROM productos ORDER BY id')
            productos = self.cursor.fetchall()
            
            if not productos:
                # Si no hay productos, asegurar que la secuencia esté en 0
                self.cursor.execute('''
                    INSERT OR REPLACE INTO sqlite_sequence (name, seq) VALUES ('productos', 0)
                ''')
                if auto_commit:
                    self.conn.commit()
                return
            
            # Verificar si los IDs ya están secuenciales
            ids_esperados = list(range(1, len(productos) + 1))
            ids_actuales = [producto[0] for producto in productos]
            
            if ids_actuales == ids_esperados:
                # Los IDs ya están correctos, solo actualizar secuencia
                self.cursor.execute(f'''
                    INSERT OR REPLACE INTO sqlite_sequence (name, seq) VALUES ('productos', {len(productos)})
                ''')
                if auto_commit:
                    self.conn.commit()
                return
            
            # Crear una tabla temporal
            self.cursor.execute('''
                CREATE TEMP TABLE productos_temp AS 
                SELECT * FROM productos WHERE 1=0
            ''')
            
            # Insertar productos con nuevos IDs secuenciales
            for nuevo_id, producto in enumerate(productos, 1):
                self.cursor.execute('''
                    INSERT INTO productos_temp (id, nombre, precio, costo, stock, categoria, codigo_barras)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (nuevo_id, producto[1], producto[2], producto[3], producto[4], producto[5], producto[6]))
            
            # Reemplazar tabla original
            self.cursor.execute('DELETE FROM productos')
            self.cursor.execute('''
                INSERT INTO productos (id, nombre, precio, costo, stock, categoria, codigo_barras)
                SELECT id, nombre, precio, costo, stock, categoria, codigo_barras 
                FROM productos_temp
            ''')
            
            # Limpiar tabla temporal
            self.cursor.execute('DROP TABLE productos_temp')
            
            # Actualizar el contador autoincrement
            self.cursor.execute(f'''
                INSERT OR REPLACE INTO sqlite_sequence (name, seq) VALUES ('productos', {len(productos)})
            ''')
            
            if auto_commit:
                self.conn.commit()
                
            print(f"IDs reorganizados: {len(productos)} productos renumerados correctamente")
            
        except Exception as e:
            print(f"Error al reorganizar IDs: {e}")
            # En caso de error, hacer rollback solo si auto_commit está habilitado
            if auto_commit:
                self.conn.rollback()
    
    def limpiar_codigos_barras(self):
        """Limpia los códigos de barras que terminan en .0 en la base de datos"""
        try:
            # Obtener todos los productos con códigos de barras que terminan en .0
            self.cursor.execute('''
                SELECT id, codigo_barras FROM productos 
                WHERE codigo_barras LIKE '%.0'
            ''')
            productos_con_problema = self.cursor.fetchall()
            
            if not productos_con_problema:
                return
            
            # Actualizar cada producto
            for producto_id, codigo_barras in productos_con_problema:
                nuevo_codigo = codigo_barras[:-2]  # Remover los últimos 2 caracteres (.0)
                self.cursor.execute('''
                    UPDATE productos SET codigo_barras = ? WHERE id = ?
                ''', (nuevo_codigo, producto_id))
            
            self.conn.commit()
            print(f"Limpiados {len(productos_con_problema)} códigos de barras")
            
        except Exception as e:
            print(f"Error al limpiar códigos de barras: {e}")
            self.conn.rollback()
    
    def confirmar_eliminar_productos(self):
        """Confirma la eliminación de todos los productos"""
        # Primera confirmación
        mensaje_1 = ("⚠️ ELIMINAR TODOS LOS PRODUCTOS\n\n"
                    "🚨 ADVERTENCIA CRÍTICA 🚨\n\n"
                    "Esta acción eliminará PERMANENTEMENTE:\n"
                    "• Todos los productos registrados\n"
                    "• Todo el inventario actual\n"
                    "• Códigos de barras asociados\n"
                    "• Información de precios y costos\n\n"
                    "❌ NO HAY FORMA DE RECUPERAR ESTA INFORMACIÓN\n\n"
                    "¿Estás COMPLETAMENTE SEGURO de continuar?")
        
        if not messagebox.askyesno("⚠️ CONFIRMACIÓN CRÍTICA", mensaje_1, icon='warning'):
            return
        
        # Segunda confirmación más estricta
        mensaje_2 = ("🔴 ÚLTIMA CONFIRMACIÓN 🔴\n\n"
                    "Vas a BORRAR PERMANENTEMENTE todos los productos.\n\n"
                    "Esta acción:\n"
                    "• NO se puede deshacer\n"
                    "• Eliminará TODO el inventario\n"
                    "• Reiniciará el catálogo a CERO\n"
                    "• Afectará el punto de venta\n\n"
                    "Para confirmar, debes presionar 'Sí' nuevamente.\n\n"
                    "¿CONFIRMAS la eliminación DEFINITIVA?")
        
        if not messagebox.askyesno("🚨 CONFIRMACIÓN FINAL", mensaje_2, icon='error'):
            return
        
        # Tercera confirmación con contraseña
        from tkinter import simpledialog
        password = simpledialog.askstring(
            "Verificación de Seguridad", 
            "Por seguridad, ingresa la contraseña del administrador\npara confirmar esta acción IRREVERSIBLE:",
            show='*'
        )
        
        if not password:
            messagebox.showinfo("Cancelado", "Operación cancelada por el usuario")
            return
        
        # Verificar contraseña del usuario actual
        if self.usuario_actual['rol'] != 'admin':
            messagebox.showerror("Sin Permisos", "Solo los administradores pueden eliminar todos los productos")
            return
        
        # Verificar contraseña en la base de datos
        self.cursor.execute(
            "SELECT password FROM usuarios WHERE nombre = ? AND rol = 'admin'",
            (self.usuario_actual['nombre'],)
        )
        user_data = self.cursor.fetchone()
        
        if not user_data or user_data[0] != password:
            messagebox.showerror("Contraseña Incorrecta", "Contraseña incorrecta. Operación cancelada.")
            return
        
        # Si llegamos aquí, proceder con la eliminación
        self.eliminar_productos()
    
    def eliminar_productos(self):
        """Elimina todos los productos de la base de datos"""
        try:
            # Contar productos antes de eliminar
            self.cursor.execute("SELECT COUNT(*) FROM productos")
            productos_count = self.cursor.fetchone()[0]
            
            # Eliminar todos los productos
            self.cursor.execute("DELETE FROM productos")
            productos_eliminados = self.cursor.rowcount
            
            # Reiniciar el contador de autoincrement
            self.cursor.execute("DELETE FROM sqlite_sequence WHERE name='productos'")
            
            # Confirmar cambios
            self.conn.commit()
            
            # Actualizar interfaces
            self.actualizar_tabla_productos()
            if hasattr(self, 'actualizar_lista_productos'):
                self.actualizar_lista_productos()
            
            # Mensaje de confirmación
            messagebox.showinfo(
                "Productos Eliminados", 
                f"✅ Eliminación completada exitosamente\n\n"
                f"• Productos eliminados: {productos_eliminados}\n"
                f"• Inventario reiniciado a cero\n"
                f"• IDs reiniciados desde 1\n\n"
                f"El catálogo de productos está ahora vacío."
            )
            
        except Exception as e:
            # Revertir cambios en caso de error
            self.conn.rollback()
            messagebox.showerror("Error", f"Error al eliminar productos: {str(e)}")
    
    # ===== MÉTODOS DE REPORTES =====
    
    def actualizar_estadisticas(self):
        """Actualiza las estadísticas de ventas"""
        hoy = datetime.now().strftime('%Y-%m-%d')
        mes_actual = datetime.now().strftime('%Y-%m')
        anio_actual = datetime.now().strftime('%Y')
        
        # Ventas de hoy
        self.cursor.execute('''
            SELECT COUNT(*), COALESCE(SUM(total), 0), COALESCE(SUM(total - costo_total), 0)
            FROM ventas WHERE DATE(fecha) = ?
        ''', (hoy,))
        ventas_hoy = self.cursor.fetchone()
        
        # Ventas del mes
        self.cursor.execute('''
            SELECT COUNT(*), COALESCE(SUM(total), 0), COALESCE(SUM(total - costo_total), 0)
            FROM ventas WHERE strftime('%Y-%m', fecha) = ?
        ''', (mes_actual,))
        ventas_mes = self.cursor.fetchone()
        
        # Ventas del año
        self.cursor.execute('''
            SELECT COUNT(*), COALESCE(SUM(total), 0), COALESCE(SUM(total - costo_total), 0)
            FROM ventas WHERE strftime('%Y', fecha) = ?
        ''', (anio_actual,))
        ventas_anio = self.cursor.fetchone()
        
        # Actualizar labels
        self.label_hoy_total.config(text=f"${ventas_hoy[1]:.2f}")
        self.label_hoy_ventas.config(text=f"{ventas_hoy[0]} ventas")
        self.label_hoy_ganancia.config(text=f"Ganancia: ${ventas_hoy[2]:.2f}")
        
        self.label_mes_total.config(text=f"${ventas_mes[1]:.2f}")
        self.label_mes_ventas.config(text=f"{ventas_mes[0]} ventas")
        self.label_mes_ganancia.config(text="Ganancia: ${:.2f}".format(ventas_mes[2]))
        
        self.label_anio_total.config(text=f"${ventas_anio[1]:.2f}")
        self.label_anio_ventas.config(text=f"{ventas_anio[0]} ventas")
        self.label_anio_ganancia.config(text="Ganancia: ${:.2f}".format(ventas_anio[2]))
        
        self.actualizar_tabla_ventas()
    
    def actualizar_tabla_ventas(self):
        """Actualiza la tabla de ventas con filtro por turno y fecha"""
        for item in self.tabla_ventas.get_children():
            self.tabla_ventas.delete(item)
        
        fecha = self.entry_fecha_reporte.get()
        turno = self.combo_turno_reporte.get()
        query = 'SELECT * FROM ventas'
        params = []
        where = []
        if fecha:
            where.append('DATE(fecha) = ?')
            params.append(fecha)
        if turno:
            where.append('turno = ?')
            params.append(turno)
        if where:
            query += ' WHERE ' + ' AND '.join(where)
        query += ' ORDER BY fecha DESC LIMIT 100'
        self.cursor.execute(query, params)
        ventas = self.cursor.fetchall()
        for venta in ventas:
            ganancia = venta[4] - venta[5]
            valores = (venta[0], venta[1], venta[2], venta[3], f'${venta[4]:.2f}', f'${ganancia:.2f}', venta[6])
            self.tabla_ventas.insert('', 'end', values=valores)
    
    def exportar_todo_excel(self):
        """Exporta todos los datos a Excel con formato profesional"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"reporte_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if filename:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Productos
                    self.cursor.execute('SELECT id, nombre, precio, costo, stock, categoria, codigo_barras FROM productos ORDER BY nombre')
                    productos = self.cursor.fetchall()
                    df_productos = pd.DataFrame(productos, columns=['ID', 'Nombre', 'Precio', 'Costo', 'Stock', 'Categoría', 'Código Barras'])
                    df_productos['Margen (%)'] = ((df_productos['Precio'] - df_productos['Costo']) / df_productos['Precio'] * 100).round(1)
                    df_productos.to_excel(writer, sheet_name='Productos', index=False)
                    
                    # Formatear hoja productos
                    ws_productos = writer.sheets['Productos']
                    self._formatear_hoja_excel(ws_productos, df_productos, 'Listado Completo de Productos')
                    
                    # Ventas con ganancia calculada
                    self.cursor.execute('SELECT * FROM ventas ORDER BY fecha DESC')
                    ventas = self.cursor.fetchall()
                    df_ventas = pd.DataFrame(ventas, columns=['ID', 'Fecha', 'Usuario', 'Método Pago', 'Total', 'Costo Total', 'Turno'])
                    df_ventas['Ganancia'] = df_ventas['Total'] - df_ventas['Costo Total']
                    df_ventas['Margen (%)'] = ((df_ventas['Total'] - df_ventas['Costo Total']) / df_ventas['Total'] * 100).round(1)
                    df_ventas.to_excel(writer, sheet_name='Todas las Ventas', index=False)
                    
                    # Formatear hoja ventas
                    ws_ventas = writer.sheets['Todas las Ventas']
                    self._formatear_hoja_excel(ws_ventas, df_ventas, 'Historial Completo de Ventas')
                    
                    # Items de ventas con más detalle
                    self.cursor.execute('''
                        SELECT iv.*, v.fecha, v.usuario 
                        FROM items_venta iv 
                        JOIN ventas v ON iv.venta_id = v.id 
                        ORDER BY v.fecha DESC
                    ''')
                    items = self.cursor.fetchall()
                    df_items = pd.DataFrame(items, columns=[
                        'ID Item', 'ID Venta', 'Producto', 'Cantidad', 'Precio Unit.', 'Costo Unit.', 'Fecha Venta', 'Usuario'
                    ])
                    df_items['Subtotal'] = df_items['Cantidad'] * df_items['Precio Unit.']
                    df_items['Ganancia Item'] = (df_items['Precio Unit.'] - df_items['Costo Unit.']) * df_items['Cantidad']
                    df_items.to_excel(writer, sheet_name='Detalle Completo', index=False)
                    
                    # Formatear hoja items
                    ws_items = writer.sheets['Detalle Completo']
                    self._formatear_hoja_excel(ws_items, df_items, 'Detalle Completo de Productos Vendidos')
                    
                    # Resumen estadístico mejorado
                    hoy = datetime.now().strftime('%Y-%m-%d')
                    mes_actual = datetime.now().strftime('%Y-%m')
                    anio_actual = datetime.now().strftime('%Y')
                    
                    self.cursor.execute('SELECT COUNT(*), SUM(total), SUM(total - costo_total) FROM ventas WHERE DATE(fecha) = ?', (hoy,))
                    resumen_hoy = self.cursor.fetchone()
                    
                    self.cursor.execute('SELECT COUNT(*), SUM(total), SUM(total - costo_total) FROM ventas WHERE strftime("%Y-%m", fecha) = ?', (mes_actual,))
                    resumen_mes = self.cursor.fetchone()
                    
                    self.cursor.execute('SELECT COUNT(*), SUM(total), SUM(total - costo_total) FROM ventas WHERE strftime("%Y", fecha) = ?', (anio_actual,))
                    resumen_anio = self.cursor.fetchone()
                    
                    # Estadísticas adicionales
                    self.cursor.execute('SELECT COUNT(*) FROM productos')
                    total_productos = self.cursor.fetchone()[0]
                    
                    self.cursor.execute('SELECT SUM(stock * costo) FROM productos')
                    valor_inventario = self.cursor.fetchone()[0] or 0
                    
                    df_resumen = pd.DataFrame({
                        'Período': ['Hoy', 'Este Mes', 'Este Año', 'Productos Registrados', 'Valor Inventario'],
                        'Cantidad': [resumen_hoy[0] or 0, resumen_mes[0] or 0, resumen_anio[0] or 0, total_productos, ''],
                        'Total Vendido ($)': [resumen_hoy[1] or 0, resumen_mes[1] or 0, resumen_anio[1] or 0, '', f"${valor_inventario:,.2f}"],
                        'Ganancia ($)': [resumen_hoy[2] or 0, resumen_mes[2] or 0, resumen_anio[2] or 0, '', '']
                    })
                    df_resumen.to_excel(writer, sheet_name='Resumen Estadístico', index=False)
                    
                    # Formatear hoja resumen
                    ws_resumen = writer.sheets['Resumen Estadístico']
                    self._formatear_hoja_excel(ws_resumen, df_resumen, 'Resumen Estadístico General')
            
            messagebox.showinfo("Éxito", f"Reporte completo exportado con formato mejorado a:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    def exportar_ventas_dia_excel(self):
        """Exporta las ventas de un día específico a Excel con formato mejorado y detalles completos"""
        fecha = self.entry_fecha_reporte.get()
        if not fecha:
            messagebox.showwarning("Fecha requerida", "Por favor ingresa una fecha en formato YYYY-MM-DD")
            return

        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"ventas_detalladas_{fecha.replace('-', '')}.xlsx"
            )
            if filename:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Consulta mejorada con detalles completos de ventas
                self.cursor.execute('''
                    SELECT 
                        v.id as venta_id,
                        v.fecha,
                        v.usuario,
                        v.metodo_pago,
                        v.total as total_venta,
                        v.costo_total as costo_venta,
                        (v.total - v.costo_total) as ganancia_venta,
                        v.turno
                    FROM ventas v 
                    WHERE DATE(v.fecha) = ?
                    ORDER BY v.fecha DESC
                ''', (fecha,))
                ventas = self.cursor.fetchall()
                
                # Consulta detallada de productos vendidos
                self.cursor.execute('''
                    SELECT 
                        v.id as venta_id,
                        v.fecha,
                        v.usuario,
                        v.turno,
                        iv.producto_nombre as producto,
                        iv.cantidad,
                        iv.precio_unitario,
                        iv.costo_unitario,
                        (iv.cantidad * iv.precio_unitario) as subtotal,
                        (iv.cantidad * iv.costo_unitario) as costo_subtotal,
                        ((iv.cantidad * iv.precio_unitario) - (iv.cantidad * iv.costo_unitario)) as ganancia_producto
                    FROM ventas v
                    JOIN items_venta iv ON v.id = iv.venta_id
                    WHERE DATE(v.fecha) = ?
                    ORDER BY v.fecha DESC, iv.producto_nombre
                ''', (fecha,))
                detalle_productos = self.cursor.fetchall()

                # Crear workbook con formato
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    
                    # Hoja 1: Resumen de Ventas
                    if ventas:
                        df_ventas = pd.DataFrame(ventas, columns=[
                            'ID Venta', 'Fecha y Hora', 'Usuario', 'Método Pago', 
                            'Total Venta', 'Costo Venta', 'Ganancia Venta', 'Turno'
                        ])
                        df_ventas.to_excel(writer, sheet_name='Resumen Ventas', index=False)
                        
                        # Formatear hoja de ventas
                        ws_ventas = writer.sheets['Resumen Ventas']
                        self._formatear_hoja_excel(ws_ventas, df_ventas, 'Resumen de Ventas del ' + fecha)
                    
                    # Hoja 2: Detalle de Productos Vendidos
                    if detalle_productos:
                        df_detalle = pd.DataFrame(detalle_productos, columns=[
                            'ID Venta', 'Fecha y Hora', 'Usuario', 'Turno', 'Producto',
                            'Cantidad', 'Precio Unit.', 'Costo Unit.', 'Subtotal', 
                            'Costo Subtotal', 'Ganancia por Producto'
                        ])
                        df_detalle.to_excel(writer, sheet_name='Detalle Productos', index=False)
                        
                        # Formatear hoja de detalle
                        ws_detalle = writer.sheets['Detalle Productos']
                        self._formatear_hoja_excel(ws_detalle, df_detalle, 'Detalle de Productos Vendidos - ' + fecha)
                    
                    # Hoja 3: Análisis y Totales
                    totales_data = self._calcular_totales_dia(fecha, ventas, detalle_productos)
                    df_totales = pd.DataFrame(totales_data)
                    df_totales.to_excel(writer, sheet_name='Análisis y Totales', index=False)
                    
                    # Formatear hoja de totales
                    ws_totales = writer.sheets['Análisis y Totales']
                    self._formatear_hoja_totales(ws_totales, df_totales, 'Análisis Financiero - ' + fecha)

                messagebox.showinfo("Éxito", f"Reporte detallado del {fecha} exportado a:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def obtener_meses_con_ventas(self):
        """Obtiene la lista de meses que tienen ventas registradas"""
        try:
            self.cursor.execute('''
                SELECT DISTINCT strftime('%Y-%m', fecha) as mes
                FROM ventas 
                ORDER BY mes DESC
            ''')
            meses = self.cursor.fetchall()
            return [mes[0] for mes in meses]
        except Exception as e:
            print(f"Error al obtener meses con ventas: {e}")
            return []
    
    def seleccionar_mes_ventas(self):
        """Muestra ventana para seleccionar mes de ventas"""
        meses_disponibles = self.obtener_meses_con_ventas()
        
        if not meses_disponibles:
            messagebox.showinfo("Sin ventas", "No hay ventas registradas para generar reportes mensuales.")
            return None
        
        # Crear ventana de selección
        ventana_mes = tk.Toplevel(self.root)
        ventana_mes.title("Seleccionar Mes para Reporte")
        ventana_mes.geometry("500x450")  # Tamaño más grande
        ventana_mes.minsize(500, 450)    # Tamaño mínimo para asegurar visibilidad
        ventana_mes.configure(bg='#FAF2E3')
        ventana_mes.grab_set()  # Hacer modal
        ventana_mes.resizable(True, True)  # Permitir redimensionar
        
        # Centrar ventana
        ventana_mes.transient(self.root)
        
        # Centrar ventana en pantalla
        ventana_mes.update_idletasks()
        width = ventana_mes.winfo_width()
        height = ventana_mes.winfo_height()
        x = (ventana_mes.winfo_screenwidth() // 2) - (width // 2)
        y = (ventana_mes.winfo_screenheight() // 2) - (height // 2)
        ventana_mes.geometry(f'{width}x{height}+{x}+{y}')
        
        mes_seleccionado = tk.StringVar()
        
        # Título
        tk.Label(
            ventana_mes,
            text="Seleccionar Mes para Reporte de Ventas",
            font=('Arial', 16, 'bold'),
            bg='#FAF2E3',
            fg='#2563eb'
        ).pack(pady=(20, 10))
        
        # Instrucciones
        tk.Label(
            ventana_mes,
            text="Elige el mes del cual deseas generar el reporte en Excel:",
            font=('Arial', 11),
            bg='#FAF2E3'
        ).pack(pady=(0, 15))
        
        # Frame para lista de meses con altura fija
        frame_lista = tk.Frame(ventana_mes, bg='#FAF2E3', height=250)
        frame_lista.pack(fill='both', expand=True, padx=20, pady=(0, 15))
        frame_lista.pack_propagate(False)  # Mantener altura fija
        
        # Scrollbar para la lista
        scrollbar = tk.Scrollbar(frame_lista)
        scrollbar.pack(side='right', fill='y')
        
        # Listbox con meses disponibles
        listbox_meses = tk.Listbox(
            frame_lista,
            font=('Arial', 11),
            yscrollcommand=scrollbar.set,
            selectmode='single'
        )
        
        # Agregar meses con formato mejorado
        for mes in meses_disponibles:
            # Convertir formato YYYY-MM a texto más legible
            try:
                fecha_obj = datetime.strptime(mes, '%Y-%m')
                texto_mes = fecha_obj.strftime('%B %Y')  # Ejemplo: "Noviembre 2025"
                # Traducir meses al español
                meses_espanol = {
                    'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
                    'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
                    'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
                    'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
                }
                for eng, esp in meses_espanol.items():
                    texto_mes = texto_mes.replace(eng, esp)
                
                listbox_meses.insert(tk.END, f"{texto_mes} ({mes})")
            except:
                listbox_meses.insert(tk.END, mes)
        
        listbox_meses.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox_meses.yview)
        
        # Seleccionar primer item por defecto
        if meses_disponibles:
            listbox_meses.selection_set(0)
        
        # Frame para botones con altura fija
        frame_botones = tk.Frame(ventana_mes, bg='#FAF2E3', height=60)
        frame_botones.pack(fill='x', padx=20, pady=(15, 20))
        frame_botones.pack_propagate(False)  # Mantener altura fija
        
        def generar_reporte():
            seleccion = listbox_meses.curselection()
            if seleccion:
                indice = seleccion[0]
                mes_elegido = meses_disponibles[indice]
                mes_seleccionado.set(mes_elegido)
                ventana_mes.destroy()
        
        def cancelar():
            mes_seleccionado.set("")
            ventana_mes.destroy()
        
        # Botones con mejor diseño y espaciado
        tk.Button(
            frame_botones,
            text="Generar Reporte",
            font=('Arial', 12, 'bold'),
            bg='#16a34a',
            fg='white',
            command=generar_reporte,
            cursor='hand2',
            width=18,
            height=2
        ).pack(side='left', padx=(0, 10), pady=10)
        
        tk.Button(
            frame_botones,
            text="Cancelar",
            font=('Arial', 12),
            bg='#dc2626',
            fg='white',
            command=cancelar,
            cursor='hand2',
            width=18,
            height=2
        ).pack(side='right', padx=(10, 0), pady=10)
        
        # Esperar a que se cierre la ventana
        ventana_mes.wait_window()
        
        return mes_seleccionado.get() if mes_seleccionado.get() else None
    
    def exportar_ventas_mes_excel(self):
        """Exporta las ventas de un mes específico a Excel con análisis completo"""
        mes_seleccionado = self.seleccionar_mes_ventas()
        
        if not mes_seleccionado:
            return
        
        try:
            # Formato de nombre de archivo más descriptivo
            fecha_obj = datetime.strptime(mes_seleccionado, '%Y-%m')
            nombre_mes = fecha_obj.strftime('%B_%Y').replace('January', 'Enero').replace('February', 'Febrero').replace('March', 'Marzo').replace('April', 'Abril').replace('May', 'Mayo').replace('June', 'Junio').replace('July', 'Julio').replace('August', 'Agosto').replace('September', 'Septiembre').replace('October', 'Octubre').replace('November', 'Noviembre').replace('December', 'Diciembre')
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"reporte_mensual_{nombre_mes}.xlsx"
            )
            
            if filename:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Consulta de ventas del mes
                self.cursor.execute('''
                    SELECT 
                        v.id as venta_id,
                        v.fecha,
                        v.usuario,
                        v.metodo_pago,
                        v.total as total_venta,
                        v.costo_total as costo_venta,
                        (v.total - v.costo_total) as ganancia_venta,
                        v.turno
                    FROM ventas v 
                    WHERE strftime('%Y-%m', v.fecha) = ?
                    ORDER BY v.fecha DESC
                ''', (mes_seleccionado,))
                ventas_mes = self.cursor.fetchall()
                
                # Consulta detallada de productos vendidos en el mes
                self.cursor.execute('''
                    SELECT 
                        v.id as venta_id,
                        v.fecha,
                        v.usuario,
                        v.turno,
                        iv.producto_nombre as producto,
                        iv.cantidad,
                        iv.precio_unitario,
                        iv.costo_unitario,
                        (iv.cantidad * iv.precio_unitario) as subtotal,
                        (iv.cantidad * iv.costo_unitario) as costo_subtotal,
                        ((iv.cantidad * iv.precio_unitario) - (iv.cantidad * iv.costo_unitario)) as ganancia_producto
                    FROM ventas v
                    JOIN items_venta iv ON v.id = iv.venta_id
                    WHERE strftime('%Y-%m', v.fecha) = ?
                    ORDER BY v.fecha DESC, iv.producto_nombre
                ''', (mes_seleccionado,))
                detalle_productos_mes = self.cursor.fetchall()
                
                # Crear workbook con formato
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    
                    # Hoja 1: Resumen de Ventas del Mes
                    if ventas_mes:
                        df_ventas = pd.DataFrame(ventas_mes, columns=[
                            'ID Venta', 'Fecha y Hora', 'Usuario', 'Método Pago', 
                            'Total Venta', 'Costo Venta', 'Ganancia Venta', 'Turno'
                        ])
                        df_ventas.to_excel(writer, sheet_name='Resumen Ventas Mes', index=False)
                        
                        # Formatear hoja de ventas
                        ws_ventas = writer.sheets['Resumen Ventas Mes']
                        self._formatear_hoja_excel(ws_ventas, df_ventas, f'Resumen de Ventas - {nombre_mes.replace("_", " ")}')
                    
                    # Hoja 2: Detalle de Productos Vendidos
                    if detalle_productos_mes:
                        df_detalle = pd.DataFrame(detalle_productos_mes, columns=[
                            'ID Venta', 'Fecha y Hora', 'Usuario', 'Turno', 'Producto',
                            'Cantidad', 'Precio Unit.', 'Costo Unit.', 'Subtotal', 
                            'Costo Subtotal', 'Ganancia por Producto'
                        ])
                        df_detalle.to_excel(writer, sheet_name='Detalle Productos Mes', index=False)
                        
                        # Formatear hoja de detalle
                        ws_detalle = writer.sheets['Detalle Productos Mes']
                        self._formatear_hoja_excel(ws_detalle, df_detalle, f'Detalle de Productos - {nombre_mes.replace("_", " ")}')
                    
                    # Hoja 3: Análisis Mensual
                    analisis_data = self._calcular_analisis_mensual(mes_seleccionado, ventas_mes, detalle_productos_mes)
                    df_analisis = pd.DataFrame(analisis_data)
                    df_analisis.to_excel(writer, sheet_name='Análisis Mensual', index=False)
                    
                    # Formatear hoja de análisis
                    ws_analisis = writer.sheets['Análisis Mensual']
                    self._formatear_hoja_totales(ws_analisis, df_analisis, f'Análisis Financiero Mensual - {nombre_mes.replace("_", " ")}')
                    
                    # Hoja 4: Resumen por Días
                    resumen_dias = self._calcular_resumen_por_dias(mes_seleccionado)
                    if resumen_dias:
                        df_dias = pd.DataFrame(resumen_dias)
                        df_dias.to_excel(writer, sheet_name='Resumen por Días', index=False)
                        
                        # Formatear hoja de resumen diario
                        ws_dias = writer.sheets['Resumen por Días']
                        self._formatear_hoja_excel(ws_dias, df_dias, f'Resumen Diario - {nombre_mes.replace("_", " ")}')

                messagebox.showinfo("Éxito", f"Reporte mensual de {nombre_mes.replace('_', ' ')} exportado exitosamente a:\n{filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar reporte mensual: {str(e)}")
    
    def _calcular_analisis_mensual(self, mes, ventas_mes, detalle_productos_mes):
        """Calcula análisis completo del mes"""
        if not ventas_mes:
            return [{'Concepto': 'Sin ventas registradas en este mes', 'Valor': 'N/A'}]
        
        # Convertir mes a formato legible
        try:
            fecha_obj = datetime.strptime(mes, '%Y-%m')
            nombre_mes = fecha_obj.strftime('%B %Y')
            # Traducir al español
            meses_espanol = {
                'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
                'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
                'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
                'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
            }
            for eng, esp in meses_espanol.items():
                nombre_mes = nombre_mes.replace(eng, esp)
        except:
            nombre_mes = mes
        
        # Cálculos básicos
        total_ventas = len(ventas_mes)
        total_vendido = sum(v[4] for v in ventas_mes)  # total_venta
        total_costo = sum(v[5] for v in ventas_mes)    # costo_venta
        total_ganancia = total_vendido - total_costo
        
        # Análisis por días únicos
        dias_con_ventas = len(set(v[1][:10] for v in ventas_mes))  # fecha (solo día)
        
        # Análisis por turno
        ventas_por_turno = {}
        for venta in ventas_mes:
            turno = venta[7]  # turno
            if turno in ventas_por_turno:
                ventas_por_turno[turno]['cantidad'] += 1
                ventas_por_turno[turno]['total'] += venta[4]
                ventas_por_turno[turno]['ganancia'] += venta[6]
            else:
                ventas_por_turno[turno] = {
                    'cantidad': 1, 
                    'total': venta[4], 
                    'ganancia': venta[6]
                }
        
        # Análisis por producto
        productos_mes = {}
        for item in detalle_productos_mes:
            producto = item[4]  # producto
            cantidad = item[5]  # cantidad
            ganancia = item[10] # ganancia_producto
            subtotal = item[8]  # subtotal
            
            if producto in productos_mes:
                productos_mes[producto]['cantidad'] += cantidad
                productos_mes[producto]['ganancia'] += ganancia
                productos_mes[producto]['ventas'] += subtotal
            else:
                productos_mes[producto] = {
                    'cantidad': cantidad, 
                    'ganancia': ganancia,
                    'ventas': subtotal
                }
        
        # Productos destacados
        if productos_mes:
            producto_mas_vendido = max(productos_mes.items(), key=lambda x: x[1]['cantidad'])
            producto_mayor_ganancia = max(productos_mes.items(), key=lambda x: x[1]['ganancia'])
            producto_mayor_ventas = max(productos_mes.items(), key=lambda x: x[1]['ventas'])
        else:
            producto_mas_vendido = ('N/A', {'cantidad': 0})
            producto_mayor_ganancia = ('N/A', {'ganancia': 0})
            producto_mayor_ventas = ('N/A', {'ventas': 0})
        
        # Turno más productivo
        turno_mas_productivo = max(ventas_por_turno.items(), key=lambda x: x[1]['total']) if ventas_por_turno else ('N/A', {'total': 0, 'cantidad': 0})
        
        # Promedios
        promedio_venta = total_vendido / total_ventas if total_ventas > 0 else 0
        promedio_dia = total_vendido / dias_con_ventas if dias_con_ventas > 0 else 0
        margen_ganancia = (total_ganancia / total_vendido * 100) if total_vendido > 0 else 0
        
        # Crear lista de análisis
        analisis = [
            {'Concepto': 'Período del Reporte', 'Valor': nombre_mes},
            {'Concepto': 'Total de Ventas Realizadas', 'Valor': total_ventas},
            {'Concepto': 'Días con Ventas', 'Valor': dias_con_ventas},
            {'Concepto': 'Total Vendido ($)', 'Valor': f"${total_vendido:,.2f}"},
            {'Concepto': 'Total Costo de Mercadería ($)', 'Valor': f"${total_costo:,.2f}"},
            {'Concepto': 'Ganancia Total del Mes ($)', 'Valor': f"${total_ganancia:,.2f}"},
            {'Concepto': 'Promedio por Venta ($)', 'Valor': f"${promedio_venta:,.2f}"},
            {'Concepto': 'Promedio por Día ($)', 'Valor': f"${promedio_dia:,.2f}"},
            {'Concepto': 'Margen de Ganancia (%)', 'Valor': f"{margen_ganancia:.1f}%"},
            {'Concepto': 'Productos Únicos Vendidos', 'Valor': len(productos_mes)},
            {'Concepto': 'Producto Más Vendido', 'Valor': f"{producto_mas_vendido[0]} ({producto_mas_vendido[1]['cantidad']} unidades)"},
            {'Concepto': 'Producto con Mayor Ganancia', 'Valor': f"{producto_mayor_ganancia[0]} (${producto_mayor_ganancia[1]['ganancia']:,.2f})"},
            {'Concepto': 'Producto con Mayores Ventas', 'Valor': f"{producto_mayor_ventas[0]} (${producto_mayor_ventas[1]['ventas']:,.2f})"},
            {'Concepto': 'Turno Más Productivo', 'Valor': f"{turno_mas_productivo[0]} (${turno_mas_productivo[1]['total']:,.2f} en {turno_mas_productivo[1]['cantidad']} ventas)"}
        ]
        
        # Agregar análisis por turno
        analisis.append({'Concepto': '--- ANÁLISIS POR TURNO ---', 'Valor': ''})
        for turno, datos in ventas_por_turno.items():
            analisis.append({
                'Concepto': f'Turno {turno}',
                'Valor': f"{datos['cantidad']} ventas - ${datos['total']:,.2f} - Ganancia: ${datos['ganancia']:,.2f}"
            })
        
        return analisis
    
    def _calcular_resumen_por_dias(self, mes):
        """Calcula resumen de ventas por día del mes"""
        try:
            self.cursor.execute('''
                SELECT 
                    DATE(fecha) as dia,
                    COUNT(*) as num_ventas,
                    SUM(total) as total_dia,
                    SUM(costo_total) as costo_dia,
                    SUM(total - costo_total) as ganancia_dia
                FROM ventas 
                WHERE strftime('%Y-%m', fecha) = ?
                GROUP BY DATE(fecha)
                ORDER BY dia
            ''', (mes,))
            
            resultados = self.cursor.fetchall()
            
            resumen_dias = []
            for resultado in resultados:
                dia, num_ventas, total_dia, costo_dia, ganancia_dia = resultado
                
                # Formatear fecha para mostrar día de la semana
                try:
                    fecha_obj = datetime.strptime(dia, '%Y-%m-%d')
                    dia_semana = fecha_obj.strftime('%A')
                    # Traducir días al español
                    dias_espanol = {
                        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
                    }
                    dia_semana = dias_espanol.get(dia_semana, dia_semana)
                    dia_formateado = f"{dia} ({dia_semana})"
                except:
                    dia_formateado = dia
                
                resumen_dias.append({
                    'Fecha': dia_formateado,
                    'Número de Ventas': num_ventas,
                    'Total Vendido ($)': f"${total_dia:,.2f}",
                    'Costo Total ($)': f"${costo_dia:,.2f}",
                    'Ganancia ($)': f"${ganancia_dia:,.2f}",
                    'Promedio por Venta ($)': f"${(total_dia/num_ventas):,.2f}" if num_ventas > 0 else "$0.00"
                })
            
            return resumen_dias
            
        except Exception as e:
            print(f"Error al calcular resumen por días: {e}")
            return []
    
    def _formatear_hoja_excel(self, worksheet, dataframe, titulo):
        """Aplica formato profesional a una hoja de Excel"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Colores y estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Insertar fila para título
        worksheet.insert_rows(1)
        worksheet['A1'] = titulo
        worksheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        worksheet['A1'].fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:' + chr(65 + len(dataframe.columns) - 1) + '1')
        
        # Formatear encabezados (ahora en fila 2)
        for col_num, column_title in enumerate(dataframe.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Formatear datos y ajustar columnas
        for row_num in range(3, len(dataframe) + 3):
            for col_num in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                
                # Formato especial para columnas de dinero
                column_name = dataframe.columns[col_num - 1].lower()
                if any(word in column_name for word in ['precio', 'costo', 'total', 'ganancia', 'subtotal']):
                    cell.number_format = '"$" #,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajustar ancho de columnas automáticamente
        for col_num in range(1, len(dataframe.columns) + 1):
            max_length = 0
            column_letter = chr(64 + col_num)  # A=65, B=66, etc.
            
            # Revisar todas las celdas de la columna
            for row_num in range(1, len(dataframe) + 3):  # +3 por título y header
                try:
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if hasattr(cell, 'value') and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Ajustar ancho mínimo para headers
            if col_num <= len(dataframe.columns):
                header_length = len(str(dataframe.columns[col_num - 1]))
                max_length = max(max_length, header_length)
            
            adjusted_width = min(max(max_length + 2, 12), 50)  # Mínimo 12, máximo 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _formatear_hoja_totales(self, worksheet, dataframe, titulo):
        """Formato especial para la hoja de totales y análisis"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Aplicar formato base
        self._formatear_hoja_excel(worksheet, dataframe, titulo)
        
        # Destacar totales importantes con colores
        for row_num in range(3, len(dataframe) + 3):
            concepto_cell = worksheet.cell(row=row_num, column=1)
            valor_cell = worksheet.cell(row=row_num, column=2)
            
            concepto = str(concepto_cell.value).lower()
            
            if 'total' in concepto or 'ganancia total' in concepto:
                concepto_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                concepto_cell.font = Font(bold=True, color='FFFFFF')
                valor_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                valor_cell.font = Font(bold=True)
            elif 'producto más vendido' in concepto or 'mejor' in concepto:
                concepto_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                concepto_cell.font = Font(bold=True)
                valor_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    def _calcular_totales_dia(self, fecha, ventas, detalle_productos):
        """Calcula análisis completo del día"""
        if not ventas:
            return [{'Concepto': 'Sin ventas registradas', 'Valor': 'N/A'}]
        
        # Cálculos básicos
        total_ventas = len(ventas)
        total_vendido = sum(v[4] for v in ventas)  # total_venta
        total_costo = sum(v[5] for v in ventas)    # costo_venta
        total_ganancia = total_vendido - total_costo
        
        # Análisis por producto
        productos_vendidos = {}
        for item in detalle_productos:
            producto = item[4]  # producto
            cantidad = item[5]  # cantidad
            ganancia = item[10] # ganancia_producto
            
            if producto in productos_vendidos:
                productos_vendidos[producto]['cantidad'] += cantidad
                productos_vendidos[producto]['ganancia'] += ganancia
            else:
                productos_vendidos[producto] = {'cantidad': cantidad, 'ganancia': ganancia}
        
        # Producto más vendido
        producto_mas_vendido = max(productos_vendidos.items(), key=lambda x: x[1]['cantidad']) if productos_vendidos else ('N/A', {'cantidad': 0})
        
        # Producto con mayor ganancia
        producto_mayor_ganancia = max(productos_vendidos.items(), key=lambda x: x[1]['ganancia']) if productos_vendidos else ('N/A', {'ganancia': 0})
        
        # Promedio por venta
        promedio_venta = total_vendido / total_ventas if total_ventas > 0 else 0
        
        # Margen de ganancia promedio
        margen_ganancia = (total_ganancia / total_vendido * 100) if total_vendido > 0 else 0
        
        return [
            {'Concepto': 'Fecha del Reporte', 'Valor': fecha},
            {'Concepto': 'Total de Ventas Realizadas', 'Valor': total_ventas},
            {'Concepto': 'Total Vendido ($)', 'Valor': f"${total_vendido:,.2f}"},
            {'Concepto': 'Total Costo de Mercadería ($)', 'Valor': f"${total_costo:,.2f}"},
            {'Concepto': 'Ganancia Total del Día ($)', 'Valor': f"${total_ganancia:,.2f}"},
            {'Concepto': 'Promedio por Venta ($)', 'Valor': f"${promedio_venta:,.2f}"},
            {'Concepto': 'Margen de Ganancia (%)', 'Valor': f"{margen_ganancia:.1f}%"},
            {'Concepto': 'Productos Únicos Vendidos', 'Valor': len(productos_vendidos)},
            {'Concepto': 'Producto Más Vendido', 'Valor': f"{producto_mas_vendido[0]} ({producto_mas_vendido[1]['cantidad']} unidades)"},
            {'Concepto': 'Producto con Mayor Ganancia', 'Valor': f"{producto_mayor_ganancia[0]} (${producto_mayor_ganancia[1]['ganancia']:,.2f})"}
        ]

    def confirmar_eliminar_reportes(self):
        """Confirma la eliminación de todos los reportes de ventas"""
        # Primera confirmación
        mensaje_1 = ("⚠️ ELIMINAR TODOS LOS REPORTES DE VENTAS\n\n"
                    "🚨 ADVERTENCIA CRÍTICA 🚨\n\n"
                    "Esta acción eliminará PERMANENTEMENTE:\n"
                    "• Todas las ventas registradas\n"
                    "• Todo el detalle de items vendidos\n"
                    "• Histórico completo de transacciones\n"
                    "• Estadísticas y reportes\n\n"
                    "❌ NO HAY FORMA DE RECUPERAR ESTA INFORMACIÓN\n\n"
                    "¿Estás COMPLETAMENTE SEGURO de continuar?")
        
        if not messagebox.askyesno("⚠️ CONFIRMACIÓN CRÍTICA", mensaje_1, icon='warning'):
            return
        
        # Segunda confirmación más estricta
        mensaje_2 = ("🔴 ÚLTIMA CONFIRMACIÓN 🔴\n\n"
                    "Vas a BORRAR PERMANENTEMENTE todos los reportes.\n\n"
                    "Esta acción:\n"
                    "• NO se puede deshacer\n"
                    "• Eliminará TODO el historial de ventas\n"
                    "• Reiniciará las estadísticas a CERO\n\n"
                    "Para confirmar, debes presionar 'Sí' nuevamente.\n\n"
                    "¿CONFIRMAS la eliminación DEFINITIVA?")
        
        if not messagebox.askyesno("🚨 CONFIRMACIÓN FINAL", mensaje_2, icon='error'):
            return
        
        # Tercera confirmación con contraseña
        from tkinter import simpledialog
        password = simpledialog.askstring(
            "Verificación de Seguridad", 
            "Por seguridad, ingresa la contraseña del administrador\npara confirmar esta acción IRREVERSIBLE:",
            show='*'
        )
        
        if not password:
            messagebox.showinfo("Cancelado", "Operación cancelada por el usuario")
            return
        
        # Verificar contraseña del usuario actual
        if self.usuario_actual['rol'] != 'admin':
            messagebox.showerror("Sin Permisos", "Solo los administradores pueden eliminar reportes")
            return
        
        # Verificar contraseña en la base de datos
        self.cursor.execute(
            "SELECT password FROM usuarios WHERE nombre = ? AND rol = 'admin'",
            (self.usuario_actual['nombre'],)
        )
        user_data = self.cursor.fetchone()
        
        if not user_data or user_data[0] != password:
            messagebox.showerror("Contraseña Incorrecta", "Contraseña incorrecta. Operación cancelada.")
            return
        
        # Si llegamos aquí, proceder con la eliminación
        self.eliminar_reportes()
    
    def eliminar_reportes(self):
        """Elimina todos los reportes de ventas de la base de datos"""
        try:
            # Eliminar items de venta primero (por restricción de clave foránea)
            self.cursor.execute("DELETE FROM items_venta")
            items_eliminados = self.cursor.rowcount
            
            # Eliminar ventas
            self.cursor.execute("DELETE FROM ventas")
            ventas_eliminadas = self.cursor.rowcount
            
            # Confirmar cambios
            self.conn.commit()
            
            # Actualizar interfaces
            self.actualizar_estadisticas()
            self.actualizar_tabla_ventas()
            
            # Mensaje de confirmación
            messagebox.showinfo(
                "Reportes Eliminados", 
                f"✅ Eliminación completada exitosamente\n\n"
                f"• Ventas eliminadas: {ventas_eliminadas}\n"
                f"• Items eliminados: {items_eliminados}\n\n"
                f"Todas las estadísticas han sido reiniciadas."
            )
            
        except Exception as e:
            # Revertir cambios en caso de error
            self.conn.rollback()
            messagebox.showerror("Error", f"Error al eliminar reportes: {str(e)}")
    
    # ===== MÉTODOS DE USUARIOS =====
    
    def agregar_usuario(self):
        """Agrega o actualiza un usuario"""
        nombre = self.user_nombre.get()
        password = self.user_password.get()
        rol = self.user_rol.get()
        
        if not all([nombre, password, rol]):
            messagebox.showwarning("Campos Vacíos", "Por favor completa todos los campos")
            return
        
        try:
            if hasattr(self, 'usuario_id') and self.usuario_id:
                # Actualizar usuario
                self.cursor.execute('''
                    UPDATE usuarios SET nombre=?, password=?, rol=? WHERE id=?
                ''', (nombre, password, rol, self.usuario_id))
                self.conn.commit()
                messagebox.showinfo("Éxito", "Usuario actualizado correctamente")
                self.usuario_id = None
                self.actualizar_tabla_usuarios()
            else:
                self.cursor.execute('''
                    INSERT INTO usuarios (nombre, password, rol)
                    VALUES (?, ?, ?)
                ''', (nombre, password, rol))
                messagebox.showinfo("Éxito", "Usuario agregado correctamente")
                self.conn.commit()
                
                self.user_nombre.delete(0, tk.END)
                self.user_password.delete(0, tk.END)
                self.user_rol.set('empleado')
                
                self.actualizar_tabla_usuarios()
                messagebox.showinfo("Éxito", "Usuario agregado correctamente")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Ya existe un usuario con ese nombre")
    
    def actualizar_tabla_usuarios(self):
        """Actualiza la tabla de usuarios"""
        for item in self.tabla_usuarios.get_children():
            self.tabla_usuarios.delete(item)
        
        self.cursor.execute('SELECT id, nombre, password, rol FROM usuarios')
        usuarios = self.cursor.fetchall()
        
        for usuario in usuarios:
            self.tabla_usuarios.insert('', 'end', values=usuario)
    
    def eliminar_usuario(self):
        """Elimina el usuario seleccionado"""
        seleccion = self.tabla_usuarios.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un usuario")
            return
        
        self.cursor.execute('SELECT COUNT(*) FROM usuarios')
        if self.cursor.fetchone()[0] <= 1:
            messagebox.showwarning("Advertencia", "Debe haber al menos un usuario en el sistema")
            return
        
        if messagebox.askyesno("Confirmar", "¿Seguro que deseas eliminar este usuario?"):
            item = self.tabla_usuarios.item(seleccion[0])
            usuario_id = item['values'][0]
            
            self.cursor.execute('DELETE FROM usuarios WHERE id = ?', (usuario_id,))
            self.conn.commit()
            
            self.actualizar_tabla_usuarios()
            messagebox.showinfo("Éxito", "Usuario eliminado correctamente")
    def editar_usuario(self):
        """Carga el usuario seleccionado en el formulario para editar"""
        seleccion = self.tabla_usuarios.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Por favor selecciona un usuario")
            return

        item = self.tabla_usuarios.item(seleccion[0])
        valores = item['values']

        self.usuario_id = valores[0]
        self.user_nombre.delete(0, tk.END)
        self.user_nombre.insert(0, valores[1])
        self.user_password.delete(0, tk.END)
        self.user_password.insert(0, valores[2])
        self.user_rol.set(valores[3])
    # ===== OTROS MÉTODOS =====
    
    def logout(self):
        """Cierra sesión"""
        if messagebox.askyesno("Cerrar Sesión", "¿Deseas cerrar sesión?"):
            self.usuario_actual = None
            for widget in self.root.winfo_children():
                widget.destroy()
            self.mostrar_login()
    
    def seleccionar_turno(self):
        """Solicita al empleado seleccionar el turno antes de mostrar el punto de venta"""
        turno_win = tk.Toplevel(self.root)
        turno_win.title("Seleccionar Turno")
        turno_win.geometry("400x300")
        turno_win.resizable(False, False)
        turno_win.grab_set()
        turno_win.transient(self.root)

        # Centrar ventana
        turno_win.update_idletasks()
        x = (turno_win.winfo_screenwidth() // 2) - (400 // 2)
        y = (turno_win.winfo_screenheight() // 2) - (300 // 2)
        turno_win.geometry(f"+{x}+{y}")

        # Deshabilitar botón de cerrar
        turno_win.protocol("WM_DELETE_WINDOW", lambda: None)

        tk.Label(turno_win, text="Selecciona el turno:", font=('Arial', 18, 'bold')).pack(pady=15)
        turno_var = tk.StringVar()
        turno_var.set('MAÑANA')
        for t in ['MAÑANA', 'TARDE', 'NOCHE']:
            tk.Radiobutton(turno_win, text=t, variable=turno_var, value=t, font=('Arial', 14)).pack(anchor='w', padx=80, pady=10)

        def confirmar():
            if not turno_var.get():
                messagebox.showwarning("Turno", "Debes seleccionar un turno")
                return
            self.turno_actual = turno_var.get()
            turno_win.grab_release()
            turno_win.destroy()
            self.crear_pestaña_venta()

        tk.Button(turno_win, text="Confirmar", font=('Arial', 14, 'bold'), bg='#2563eb', fg='white', command=confirmar, width=15, height=2).pack(pady=10)
    
    def __del__(self):
        """Cierra la conexión a la base de datos"""
        if hasattr(self, 'conn'):
            self.conn.close()


if __name__ == '__main__':
    root = tk.Tk()
    app = KioscoPOS(root)
    root.mainloop()